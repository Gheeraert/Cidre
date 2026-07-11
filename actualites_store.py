# actualites_store.py — logique Excel de l'éditeur d'actualités (sans interface graphique)
# © 2026 Tony Gheeraert - Licence MIT (voir LICENSE)
#
# Ce module lit et écrit UNIQUEMENT la feuille ACTUS d'un classeur Cidre.
# Il ne lit jamais la feuille CONFIG (qui peut contenir des identifiants FTP).
# Les conventions (noms de colonnes tolérés, résolution des images) reproduisent
# exactement celles de build_site.py pour que ce que l'éditeur enregistre soit
# ce que Cidre publie.

from __future__ import annotations

import dataclasses
import os
import re
import shutil
import unicodedata
from copy import copy
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Dict, List, Optional

import openpyxl


# -------------------------
# Erreurs à message "éditrice"
# -------------------------

class ActuError(Exception):
    """Erreur affichable telle quelle à l'utilisatrice."""


class WorkbookLockedError(ActuError):
    """Le classeur est ouvert dans Excel."""


# -------------------------
# Conventions partagées avec build_site.py
# -------------------------

ALLOWED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}

# Nom du dossier canonique des images d'actualités, à côté du classeur.
# build_site.resolve_actu_image_source() cherche notamment dans excel_dir/"actu"/<nom>.
IMAGES_DIRNAME = "actu"


def _slug(s: str) -> str:
    """Même normalisation que build_site.slugify (pour comparer noms de feuilles/colonnes)."""
    s = (s or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return re.sub(r"-{2,}", "-", s)


# Variantes de colonnes reconnues par build_site.load_actualites()
FIELD_ALIASES: Dict[str, set] = {
    "title": {"title", "titre"},
    "image": {"image", "visuel", "image-file", "image_file", "cover-file", "cover_file"},
    "date": {"date", "date-publication", "date_publication", "datepub"},
    "is_active": {"is-active", "is_active", "is_published", "actif", "active"},
    "text": {"texte", "text", "contenu", "content", "resume", "description"},
    "id13": {"id13", "isbn", "isbn13", "book_id13", "gtin", "ean13", "ean-13", "isbn-13"},
    "link": {"lien", "link", "url", "lien_externe", "lien-externe"},
}

ACTUS_SHEET_NAMES = {"actualites", "actus", "news"}


def normalize_id13(v) -> str:
    """ISBN/GTIN -> chaîne de 13 chiffres, sinon '' (même logique que build_site)."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    if "e+" in s.lower():
        try:
            s = str(int(float(s)))
        except Exception:
            pass
    if s.endswith(".0"):
        s = s[:-2]
    s = re.sub(r"\D", "", s)
    return s if len(s) == 13 else ""


def parse_date_fr(s: str) -> Optional[date]:
    """'JJ/MM/AAAA' -> date. Tolère aussi JJ-MM-AAAA et AAAA-MM-JJ (anciens contenus).

    Renvoie None pour une chaîne vide ; lève ValueError avec un message clair sinon.
    """
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    raise ValueError(
        f"La date « {s} » n'est pas valide. Utilisez le format JJ/MM/AAAA (ex. 31/12/2026)."
    )


def format_date_fr(d: Optional[date]) -> str:
    return d.strftime("%d/%m/%Y") if d else ""


@dataclasses.dataclass
class Actu:
    row: Optional[int] = None  # numéro de ligne Excel ; None = nouvelle actualité
    title: str = ""
    text: str = ""
    date: Optional[date] = None
    date_raw: str = ""  # valeur brute si la cellule n'est pas une vraie date
    is_active: bool = True
    id13: str = ""
    link: str = ""
    image: str = ""  # nom de fichier tel qu'attendu par Cidre


@dataclasses.dataclass
class Book:
    id13: str
    title: str
    slug: str
    cover_file: str


@dataclasses.dataclass
class Issue:
    message: str
    confirmable: bool = False  # True = l'éditrice peut passer outre après confirmation


class ActualitesStore:
    """Accès en lecture/écriture à la feuille ACTUS d'un classeur Cidre."""

    def __init__(self, path: Path | str):
        self.path = Path(path)
        if not self.path.exists():
            raise ActuError(f"Le classeur {self.path.name} est introuvable.")
        # data_only=False (défaut) : les formules des autres feuilles sont conservées.
        self.wb = openpyxl.load_workbook(self.path)
        self.sheet_name = self._detect_actus_sheet()
        if not self.sheet_name:
            raise ActuError(
                "Ce classeur ne contient pas de feuille « ACTUS ».\n"
                "Vérifiez que vous avez ouvert le bon fichier."
            )
        self.ws = self.wb[self.sheet_name]
        self.cols = self._map_columns()
        missing = [f for f in ("title", "text", "date", "image", "is_active", "id13", "link")
                   if f not in self.cols]
        if missing:
            raise ActuError(
                "La feuille ACTUS ne contient pas toutes les colonnes attendues "
                f"(manquantes : {', '.join(missing)}).\n"
                "Ne renommez pas les en-têtes de la première ligne."
            )
        self._backup_done = False
        self._books: Optional[Dict[str, Book]] = None

    # ---------- détection structure ----------

    def _detect_actus_sheet(self) -> str:
        for sh in self.wb.sheetnames:
            if _slug(sh) in ACTUS_SHEET_NAMES:
                return sh
        return ""

    def _map_columns(self) -> Dict[str, int]:
        cols: Dict[str, int] = {}
        for idx, cell in enumerate(self.ws[1], start=1):
            lc = _slug(str(cell.value)) if cell.value is not None else ""
            if not lc:
                continue
            for field, aliases in FIELD_ALIASES.items():
                if lc in aliases and field not in cols:
                    cols[field] = idx
        return cols

    # ---------- lecture ----------

    def list_actus(self) -> List[Actu]:
        out: List[Actu] = []
        for row in range(2, self.ws.max_row + 1):
            values = {f: self.ws.cell(row=row, column=c).value for f, c in self.cols.items()}
            if all(v is None or str(v).strip() == "" for v in values.values()):
                continue
            a = Actu(row=row)
            a.title = str(values["title"]).strip() if values["title"] is not None else ""
            a.text = str(values["text"]) if values["text"] is not None else ""
            a.link = str(values["link"]).strip() if values["link"] is not None else ""
            a.image = str(values["image"]).strip() if values["image"] is not None else ""
            a.id13 = normalize_id13(values["id13"])

            d = values["date"]
            if isinstance(d, datetime):
                a.date = d.date()
            elif isinstance(d, date):
                a.date = d
            elif d is not None and str(d).strip():
                try:
                    a.date = parse_date_fr(str(d).strip())
                except ValueError:
                    a.date_raw = str(d).strip()

            v = values["is_active"]
            if v is None or str(v).strip() == "":
                a.is_active = True  # même convention que Cidre : vide = actif
            else:
                a.is_active = str(v).strip().lower() in {"1", "1.0", "true", "vrai", "oui", "x", "yes", "y"}
            out.append(a)
        return out

    def first_empty_row(self) -> int:
        """Première ligne réellement vide après les données."""
        last_data = 1
        for row in range(2, self.ws.max_row + 1):
            if any(self.ws.cell(row=row, column=c).value not in (None, "")
                   for c in self.cols.values()):
                last_data = row
        return last_data + 1

    # ---------- écriture (feuille ACTUS uniquement) ----------

    def save_actu(self, actu: Actu) -> int:
        """Écrit l'actualité dans sa ligne (ou une nouvelle ligne). Ne sauvegarde pas le fichier."""
        is_new = actu.row is None
        row = self.first_empty_row() if is_new else int(actu.row)

        if is_new and row > 2:
            self._copy_row_style(row - 1, row)

        def put(field: str, value):
            self.ws.cell(row=row, column=self.cols[field]).value = value

        put("title", actu.title or None)
        put("text", actu.text or None)
        put("link", actu.link or None)
        put("image", Path(actu.image.replace("\\", "/")).name if actu.image else None)
        put("is_active", 1 if actu.is_active else 0)
        put("id13", int(actu.id13) if actu.id13 else None)

        date_cell = self.ws.cell(row=row, column=self.cols["date"])
        if actu.date:
            date_cell.value = actu.date
            date_cell.number_format = "DD/MM/YYYY"
        else:
            date_cell.value = None

        actu.row = row
        return row

    def set_active(self, row: int, active: bool) -> None:
        self.ws.cell(row=row, column=self.cols["is_active"]).value = 1 if active else 0

    def _copy_row_style(self, src_row: int, dst_row: int) -> None:
        for col in self.cols.values():
            src = self.ws.cell(row=src_row, column=col)
            dst = self.ws.cell(row=dst_row, column=col)
            try:
                dst.font = copy(src.font)
                dst.alignment = copy(src.alignment)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.number_format = src.number_format
            except Exception:
                pass  # le style est un confort, jamais bloquant

    # ---------- sauvegarde sûre du classeur ----------

    def save_workbook(self) -> None:
        """Sauvegarde horodatée (une fois), écriture temporaire, remplacement atomique."""
        lock = self.path.parent / f"~${self.path.name}"
        if lock.exists():
            raise WorkbookLockedError("Fermez le classeur dans Excel avant de l'enregistrer.")

        if not self._backup_done:
            stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
            backup = self.path.with_name(f"{self.path.stem}.sauvegarde-{stamp}{self.path.suffix}")
            shutil.copy2(self.path, backup)
            self._backup_done = True

        tmp = self.path.with_name(f"{self.path.stem}.tmp-actus{self.path.suffix}")
        try:
            self.wb.save(tmp)
            os.replace(tmp, self.path)
        except PermissionError:
            raise WorkbookLockedError("Fermez le classeur dans Excel avant de l'enregistrer.")
        finally:
            if tmp.exists():
                try:
                    tmp.unlink()
                except Exception:
                    pass

    # ---------- catalogue (Master_Site) ----------

    def _detect_books_sheet(self) -> str:
        for sh in self.wb.sheetnames:
            if _slug(sh) == "config":
                continue  # jamais de lecture de CONFIG
            header = [str(c.value) if c.value is not None else "" for c in self.wb[sh][1]]
            lset = {_slug(h) for h in header}
            if "id13" in lset and "titre-norm" in lset:
                return sh
        return ""

    def books(self) -> Dict[str, Book]:
        if self._books is not None:
            return self._books
        self._books = {}
        sh = self._detect_books_sheet()
        if not sh:
            return self._books
        ws = self.wb[sh]
        header = {_slug(str(c.value)): i for i, c in enumerate(ws[1], start=1) if c.value}
        c_id = header.get("id13")
        c_title = header.get("titre-norm")
        c_slug = header.get("slug")
        c_cover = header.get("cover-file")
        for row in ws.iter_rows(min_row=2, values_only=True):
            i13 = normalize_id13(row[c_id - 1]) if c_id else ""
            if not i13:
                continue
            self._books[i13] = Book(
                id13=i13,
                title=str(row[c_title - 1]).strip() if c_title and row[c_title - 1] else "",
                slug=str(row[c_slug - 1]).strip() if c_slug and row[c_slug - 1] else "",
                cover_file=str(row[c_cover - 1]).strip() if c_cover and row[c_cover - 1] else "",
            )
        return self._books

    def lookup_isbn(self, raw: str) -> tuple[str, Optional[Book]]:
        """Renvoie (isbn normalisé ou '', fiche livre ou None)."""
        i13 = normalize_id13(raw)
        if not i13:
            return "", None
        return i13, self.books().get(i13)

    # ---------- images ----------

    @property
    def images_dir(self) -> Path:
        return self.path.parent / IMAGES_DIRNAME

    def resolve_image(self, img: str) -> Optional[Path]:
        """Mêmes candidats que build_site.resolve_actu_image_source()."""
        if not img:
            return None
        excel_dir = self.path.parent
        rel = img.replace("\\", "/").strip()
        candidates = [
            excel_dir / rel,
            excel_dir / "assets" / rel,
            excel_dir / "actu" / rel,
            excel_dir / "images" / rel,
            excel_dir / Path(rel).name,
            excel_dir / "assets" / Path(rel).name,
        ]
        for p in candidates:
            if p.exists() and p.is_file():
                return p
        return None

    def import_image(self, src: Path | str,
                     on_conflict: Callable[[str], str]) -> Optional[str]:
        """Copie une image dans le dossier canonique et renvoie le nom à stocker.

        on_conflict(nom) est appelé si un fichier homonyme existe déjà et doit
        renvoyer 'replace', 'keep' ou 'cancel'.
        Renvoie None si l'opération est annulée.
        """
        src = Path(src)
        if not src.exists() or not src.is_file():
            raise ActuError(f"Le fichier {src.name} est introuvable.")
        if src.suffix.lower() not in ALLOWED_IMAGE_EXTS:
            exts = ", ".join(sorted(ALLOWED_IMAGE_EXTS))
            raise ActuError(
                f"Le format de {src.name} n'est pas pris en charge par le site.\n"
                f"Formats acceptés : {exts}."
            )
        self.images_dir.mkdir(parents=True, exist_ok=True)
        dst = self.images_dir / src.name
        if dst.exists():
            if src.resolve() == dst.resolve():
                return src.name  # déjà au bon endroit
            choice = on_conflict(src.name)
            if choice == "cancel":
                return None
            if choice == "keep":
                return dst.name
        shutil.copy2(src, dst)
        return dst.name

    def find_cover_source(self, cover_file: str) -> Optional[Path]:
        """Cherche le fichier de couverture dans les emplacements plausibles."""
        if not cover_file:
            return None
        name = Path(cover_file.replace("\\", "/")).name
        excel_dir = self.path.parent
        for d in (excel_dir / "covers", excel_dir / IMAGES_DIRNAME, excel_dir / "assets",
                  excel_dir / "images", excel_dir, excel_dir / "dist" / "covers"):
            p = d / name
            if p.exists() and p.is_file():
                return p
        return None

    def use_cover(self, cover_file: str,
                  on_conflict: Callable[[str], str]) -> Optional[str]:
        """Utilise la couverture d'un livre comme image d'actualité.

        Si Cidre sait déjà la trouver (déjà dans un emplacement résolu), aucun
        fichier n'est copié. Sinon elle est copiée dans le dossier canonique.
        """
        if not cover_file:
            raise ActuError("Ce livre n'a pas de couverture renseignée dans le catalogue.")
        name = Path(cover_file.replace("\\", "/")).name
        if self.resolve_image(name):
            return name  # déjà trouvable par Cidre : pas de copie inutile
        src = self.find_cover_source(name)
        if not src:
            raise ActuError(
                f"La couverture {name} n'a pas été trouvée sur cet ordinateur.\n"
                "Utilisez « Choisir une image… » pour sélectionner le fichier."
            )
        return self.import_image(src, on_conflict)

    # ---------- validations ----------

    def validate(self, actu: Actu) -> List[Issue]:
        issues: List[Issue] = []

        if not actu.title.strip() and not actu.text.strip():
            issues.append(Issue(
                "Renseignez au moins un titre ou un texte pour cette actualité."
            ))

        if actu.date_raw:
            issues.append(Issue(
                f"La date « {actu.date_raw} » n'est pas valide. "
                "Utilisez le format JJ/MM/AAAA (ex. 31/12/2026)."
            ))

        if actu.id13:
            if not re.fullmatch(r"\d{13}", actu.id13):
                issues.append(Issue(
                    f"L'ISBN « {actu.id13} » doit comporter exactement 13 chiffres."
                ))
            elif actu.id13 not in self.books():
                issues.append(Issue(
                    f"L'ISBN {actu.id13} n'existe pas dans le catalogue (Master_Site).\n"
                    "Voulez-vous enregistrer quand même ?",
                    confirmable=True,
                ))

        if actu.image:
            name = Path(actu.image.replace("\\", "/")).name
            if not self.resolve_image(name):
                issues.append(Issue(
                    f"L'image « {name} » n'a pas été trouvée.\n"
                    "Choisissez son fichier avant d'enregistrer l'actualité."
                ))

        if actu.link and not re.match(r"^https?://", actu.link.strip(), flags=re.I):
            issues.append(Issue(
                "Le lien doit commencer par http:// ou https://\n"
                f"(valeur actuelle : {actu.link})"
            ))

        return issues
