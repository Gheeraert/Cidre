import csv
import json
import sys
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import build_site as bs
from cidre import routes
from cidre.data_models import load_config
from cidre.excel_data import detect_books_sheet, load_books, load_collections, load_revues
from cidre.url_inventory import INVENTORY_COLUMNS, collect_url_inventory, main as inventory_main


BOOK_HEADERS = [
    "id13", "slug", "titre_norm", "sous_titre_norm", "credit_ligne",
    "collection", "collection_id", "date_parution_norm", "format_site",
    "price", "availability", "cover_file", "Description courte",
    "Description longue", "order_url", "openedition_url", "active_site",
]


def _book(isbn, slug, title, collection="Essais", collection_id="col-essais"):
    return [
        isbn, slug, title, "", "", collection, collection_id,
        "2026-01-01", "Broche", "12", "Disponible", "", "Resume", "",
        "", "", 1,
    ]


def _workbook(path: Path, pages_rows=None, revues_rows=None, actualites_rows=None, include_pages=True) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFIG"
    ws.append(["key", "value"])
    for key, value in [
        ("site_title", "Site test"),
        ("books_sheet", "CATALOGUE"),
        ("pages_sheet", "PAGES"),
        ("collections_sheet", "COLLECTIONS"),
        ("revues_sheet", "REVUES"),
        ("contacts_sheet", "CONTACTS"),
    ]:
        ws.append([key, value])

    books = wb.create_sheet("CATALOGUE")
    books.append(BOOK_HEADERS)
    books.append(_book("9782877750001", "mon-livre-2", "Slug deja suffixe"))
    books.append(_book("9782877750002", "", "Fallback ISBN"))
    books.append(_book("9782877750003", "meme-slug", "Premier doublon"))
    books.append(_book("9782877750004", "meme-slug", "Second doublon"))
    books.append(_book("", "", "Sans ISBN"))
    books.append(_book("", "", "Sans ISBN"))
    books.append(_book("9782877750005", "livre-actu", "Livre actu"))
    books.append(_book("9782877750006", "numero-revue", "Numero revue", "Revue test", "revue-test"))

    if include_pages:
        pages = wb.create_sheet("PAGES")
        pages.append(["slug", "title", "content_md", "is_published"])
        for row in (pages_rows if pages_rows is not None else [
            ["presentation", "Presentation", "Texte", 1],
            ["brouillon", "Brouillon", "Texte", 0],
            ["actualites", "Actualites", "Texte", 1],
        ]):
            pages.append(row)

    collections = wb.create_sheet("COLLECTIONS")
    collections.append(["collection_id", "name", "slug", "is_active"])
    collections.append(["col-essais", "Essais", "essais", 1])

    revues = wb.create_sheet("REVUES")
    revues.append(["journal_id", "title", "slug", "is_active"])
    for row in (revues_rows if revues_rows is not None else [["revue-test", "Revue test", "revue-test", 1]]):
        revues.append(row)

    actualites = wb.create_sheet("ACTUALITES")
    actualites.append(["title", "image", "date", "text", "is_active", "id13", "link"])
    for row in (actualites_rows if actualites_rows is not None else [
        ["Actu livre", "", "2026-01-01", "Texte", 1, "9782877750005", ""],
    ]):
        actualites.append(row)

    contacts = wb.create_sheet("CONTACTS")
    contacts.append(["label", "name", "role", "email", "phone", "address", "order", "is_active"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _books(path: Path):
    with pd.ExcelFile(path) as wb:
        cfg = load_config(wb, "CONFIG")
        return load_books(wb, detect_books_sheet(wb, cfg.books_sheet))


def _inventory_rows(path: Path):
    return collect_url_inventory(path)


def _rows_by(rows, entity_type):
    return [r for r in rows if r["entity_type"] == entity_type]


def test_provenance_slug_explicite_non_unicifie(tmp_path):
    books = _books(_workbook(tmp_path / "site.xlsx"))
    row = books[books["titre_norm"] == "Slug deja suffixe"].iloc[0]
    assert row["_source_slug"] == "mon-livre-2"
    assert row["_slug_candidate"] == "mon-livre-2"
    assert row["slug"] == "mon-livre-2"
    assert row["_slug_was_uniquified"] is False
    assert row["_slug_origin"] == "explicit"


def test_provenance_revues_depuis_sources_excel(tmp_path):
    wb_path = _workbook(tmp_path / "revues.xlsx", revues_rows=[
        ["jid-explicit", "Titre explicite", "revue-test", 1],
        ["jid-title", "Revue test", "", 1],
        ["jid-only", "", "", 1],
        ["", "", "", 1],
    ])
    wb = pd.ExcelFile(wb_path)
    cfg = load_config(wb, "CONFIG")
    revues = load_revues(wb, cfg.revues_sheet)

    expected = [
        ("revue-test", "Titre explicite", "explicit", "revue-test", "revue-test"),
        ("", "Revue test", "fallback_title", "revue-test", "revue-test"),
        ("", "", "fallback_journal_id", "jid-only", "jid-only"),
        ("", "", "fallback_revue", "revue", "revue"),
    ]
    for (_, row), (source_slug, source_title, origin, candidate, final_slug) in zip(revues.iterrows(), expected):
        assert row["_source_slug"] == source_slug
        assert row["_source_title"] == source_title
        assert row["_slug_origin"] == origin
        assert row["_slug_candidate"] == candidate
        assert row["slug"] == final_slug

    inventory = _rows_by(_inventory_rows(wb_path), "revue")
    assert [(r["explicit_slug"], r["slug_origin"], r["slug_candidate"], r["final_slug"]) for r in inventory] == [
        ("revue-test", "explicit", "revue-test", "revue-test"),
        ("", "fallback_title", "revue-test", "revue-test"),
        ("", "fallback_journal_id", "jid-only", "jid-only"),
        ("", "fallback_revue", "revue", "revue"),
    ]


def test_provenance_doublon_explicite(tmp_path):
    books = _books(_workbook(tmp_path / "site.xlsx"))
    first = books[books["titre_norm"] == "Premier doublon"].iloc[0]
    second = books[books["titre_norm"] == "Second doublon"].iloc[0]
    assert first["_slug_candidate"] == "meme-slug"
    assert first["slug"] == "meme-slug"
    assert first["_slug_was_uniquified"] is False
    assert first["_slug_origin"] == "explicit"
    assert second["_slug_candidate"] == "meme-slug"
    assert second["slug"] == "meme-slug-2"
    assert second["_slug_was_uniquified"] is True
    assert second["_slug_origin"] == "explicit"


def test_provenance_fallbacks(tmp_path):
    books = _books(_workbook(tmp_path / "site.xlsx"))
    isbn = books[books["titre_norm"] == "Fallback ISBN"].iloc[0]
    assert isbn["_slug_origin"] == "fallback_title_isbn"
    assert isbn["_slug_candidate"] == "fallback-isbn-9782877750002"
    assert isbn["slug"] == "fallback-isbn-9782877750002"

    sans = books[books["titre_norm"] == "Sans ISBN"]
    assert sans.iloc[0]["_slug_origin"] == "fallback_title"
    assert sans.iloc[0]["_slug_candidate"] == "sans-isbn"
    assert sans.iloc[0]["slug"] == "sans-isbn"
    assert sans.iloc[0]["_slug_was_uniquified"] is False
    assert sans.iloc[1]["_slug_candidate"] == "sans-isbn"
    assert sans.iloc[1]["slug"] == "sans-isbn-2"
    assert sans.iloc[1]["_slug_was_uniquified"] is True


def test_coherence_producteurs_consommateurs(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "dist"
    bs.build_site(wb, out, covers_dir=None, force_alerts=True)

    books = _books(wb)
    livre_actu = books[books["titre_norm"] == "Livre actu"].iloc[0]
    numero_revue = books[books["titre_norm"] == "Numero revue"].iloc[0]
    assert (out / routes.book_public_path(livre_actu["slug"])).exists()

    actualites = json.loads((out / "assets" / "actualites.json").read_text(encoding="utf-8"))
    assert actualites[0]["href"] == routes.actualite_book_href(livre_actu["slug"], ".")

    wb_xl = pd.ExcelFile(wb)
    cfg = load_config(wb_xl, "CONFIG")
    collections = load_collections(wb_xl, cfg.collections_sheet)
    revues = load_revues(wb_xl, cfg.revues_sheet)
    collection = collections.iloc[0]
    revue = revues.iloc[0]
    assert (out / routes.collection_public_path(collection["slug"], collection["collection_id"])).exists()
    assert routes.collection_href(collection["slug"], collection["collection_id"], "..") in (
        out / routes.book_public_path(livre_actu["slug"])
    ).read_text(encoding="utf-8")
    assert (out / routes.revue_public_path(revue["slug"], revue["title"], revue["journal_id"])).exists()
    assert routes.revue_href(revue["slug"], revue["title"], revue["journal_id"], "..") in (
        out / routes.book_public_path(numero_revue["slug"])
    ).read_text(encoding="utf-8")


def test_inventaire_csv(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx")
    before = wb.read_bytes()
    csv_path = tmp_path / "audit" / "url_inventory.csv"

    exit_code = inventory_main(["--excel", str(wb), "--csv", str(csv_path)])

    assert exit_code == 0
    assert wb.read_bytes() == before
    assert not (tmp_path / "dist").exists()
    rows = list(csv.DictReader(csv_path.open(encoding="utf-8")))
    assert rows
    assert list(rows[0].keys()) == INVENTORY_COLUMNS

    categories = {r["entity_type"] for r in rows}
    assert {
        "automatic_page", "book", "collection", "revue", "editorial_page",
        "actualite_anchor", "actualite_book_link",
    } <= categories

    by_title = {r["title"]: r for r in rows if r["entity_type"] == "book"}
    assert by_title["Slug deja suffixe"]["explicit_slug"] == "mon-livre-2"
    assert by_title["Slug deja suffixe"]["auto_uniquified"] == "False"
    assert by_title["Second doublon"]["slug_candidate"] == "meme-slug"
    assert by_title["Second doublon"]["final_slug"] == "meme-slug-2"
    assert by_title["Second doublon"]["auto_uniquified"] == "True"
    assert by_title["Fallback ISBN"]["slug_origin"] == "fallback_title_isbn"
    assert by_title["Sans ISBN"]["slug_origin"] == "fallback_title"
    assert any(r["public_path"] == "livres/livre-actu.html" for r in rows)
    assert any(r["public_path"] == "actualites.html#actu-actu-livre" for r in rows)


def _open_access_rows(path: Path):
    rows = collect_url_inventory(path)
    return [r for r in rows if r["public_path"] == "open-access.html"]


def test_inventaire_open_access_page_publiee(tmp_path):
    wb = _workbook(tmp_path / "open-access.xlsx", pages_rows=[
        ["open-access", "Open access", "Texte", 1],
    ])
    rows = _open_access_rows(wb)
    assert len(rows) == 1
    assert rows[0]["entity_type"] == "editorial_page"
    assert rows[0]["slug_origin"] == "explicit"


def test_inventaire_open_access_absente(tmp_path):
    wb = _workbook(tmp_path / "sans-open-access.xlsx", pages_rows=[
        ["presentation", "Presentation", "Texte", 1],
    ])
    rows = _open_access_rows(wb)
    assert len(rows) == 1
    assert rows[0]["entity_type"] == "fallback_page"
    assert rows[0]["identifier"] == "open-access"
    assert rows[0]["slug_origin"] == "generated_fallback"
    assert rows[0]["notes"] == "Page de secours garantie par build_pages"


def test_inventaire_open_access_non_publiee(tmp_path):
    wb = _workbook(tmp_path / "open-access-brouillon.xlsx", pages_rows=[
        ["open-access", "Open access", "Texte", 0],
    ])
    rows = _open_access_rows(wb)
    assert len(rows) == 1
    assert rows[0]["entity_type"] == "fallback_page"


def test_inventaire_open_access_sans_feuille_pages(tmp_path):
    wb = _workbook(tmp_path / "sans-pages.xlsx", include_pages=False)
    rows = _open_access_rows(wb)
    assert len(rows) == 1
    assert rows[0]["entity_type"] == "fallback_page"


def test_inventaire_actualites_candidat_final_et_unicisation(tmp_path):
    wb = _workbook(tmp_path / "actus.xlsx", actualites_rows=[
        ["Même titre", "", "2026-01-03", "Texte", 1, "", ""],
        ["Même titre", "", "2026-01-02", "Texte", 1, "", ""],
        ["", "", "2026-01-01", "Texte sans titre", 1, "", ""],
    ])
    anchors = _rows_by(collect_url_inventory(wb), "actualite_anchor")
    by_path = {r["public_path"]: r for r in anchors}

    first = by_path["actualites.html#actu-meme-titre"]
    assert first["slug_origin"] == "fallback_title"
    assert first["slug_candidate"] == "meme-titre"
    assert first["final_slug"] == "meme-titre"
    assert first["auto_uniquified"] == "False"

    second = by_path["actualites.html#actu-meme-titre-2"]
    assert second["slug_origin"] == "fallback_title"
    assert second["slug_candidate"] == "meme-titre"
    assert second["final_slug"] == "meme-titre-2"
    assert second["auto_uniquified"] == "True"

    fallback = by_path["actualites.html#actu-actu"]
    assert fallback["slug_origin"] == "fallback_actu"
    assert fallback["slug_candidate"] == "actu"
    assert fallback["final_slug"] == "actu"
    assert fallback["auto_uniquified"] == "False"


def test_inventaire_actualites_ponctuation_fallback_actu(tmp_path):
    wb = _workbook(tmp_path / "actus-ponctuation.xlsx", actualites_rows=[
        ["!!!", "", "2026-01-03", "Texte", 1, "", ""],
        ["!!!", "", "2026-01-02", "Texte", 1, "", ""],
        ["", "", "2026-01-01", "Texte sans titre", 1, "", ""],
    ])
    anchors = _rows_by(collect_url_inventory(wb), "actualite_anchor")
    by_path = {r["public_path"]: r for r in anchors}

    first = by_path["actualites.html#actu-actu"]
    assert first["title"] == "!!!"
    assert first["slug_origin"] == "fallback_title"
    assert first["slug_candidate"] == "actu"
    assert first["final_slug"] == "actu"
    assert first["auto_uniquified"] == "False"

    second = by_path["actualites.html#actu-actu-2"]
    assert second["title"] == "!!!"
    assert second["slug_origin"] == "fallback_title"
    assert second["slug_candidate"] == "actu"
    assert second["final_slug"] == "actu-2"
    assert second["auto_uniquified"] == "True"

    missing_title = by_path["actualites.html#actu-actu-3"]
    assert missing_title["title"] == ""
    assert missing_title["slug_origin"] == "fallback_actu"
    assert missing_title["slug_candidate"] == "actu"
