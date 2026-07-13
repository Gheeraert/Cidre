import sys
import warnings
from pathlib import Path

import openpyxl
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import build_site as bs
import cidre.orchestrator as orchestrator
import cidre.output_transaction as output_transaction


def _workbook(path: Path, title: str = "Un livre", pages_rows: list[dict] | None = None) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFIG"
    ws.append(["key", "value"])
    for key, value in [
        ("site_title", "Site test"),
        ("books_sheet", "CATALOGUE"),
        ("pages_sheet", "PAGES"),
        ("collections_sheet", "COLLECTIONS"),
        ("revues_sheet", "REVUES"),
        ("contacts_sheet", "CONTACTS"),
    ]:
        ws.append([key, value])

    books = wb.create_sheet("CATALOGUE")
    books.append([
        "id13", "slug", "titre_norm", "sous_titre_norm", "credit_ligne",
        "collection", "collection_id", "date_parution_norm", "format_site",
        "price", "availability", "cover_file", "Description courte",
        "Description longue", "order_url", "openedition_url", "active_site",
    ])
    books.append([
        "9782877750001", "un-livre", title, "", "", "Essais", "col-essais",
        "2026-01-01", "Broche", "12", "Disponible", "", "Resume", "",
        "", "", 1,
    ])

    pages = wb.create_sheet("PAGES")
    pages.append(["slug", "title", "content_md", "is_published"])
    if pages_rows is None:
        pages_rows = [{"slug": "presentation", "title": "Presentation", "content_md": "Texte neuf", "is_published": 1}]
    for row in pages_rows:
        pages.append([
            row.get("slug", ""),
            row.get("title", ""),
            row.get("content_md", ""),
            row.get("is_published", 1),
        ])

    collections = wb.create_sheet("COLLECTIONS")
    collections.append(["collection_id", "name", "slug", "is_active"])
    collections.append(["col-essais", "Essais", "essais", 1])

    revues = wb.create_sheet("REVUES")
    revues.append(["revue_id", "title", "slug", "is_active"])

    contacts = wb.create_sheet("CONTACTS")
    contacts.append(["label", "name", "role", "email", "phone", "address", "order", "is_active"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def _leftovers(out: Path) -> list[Path]:
    return sorted(out.parent.glob(f".{out.name}.build-*")) + sorted(out.parent.glob(f".{out.name}.backup-*"))


def _snapshot(path: Path) -> dict[str, bytes]:
    return {
        p.relative_to(path).as_posix(): p.read_bytes()
        for p in sorted(path.rglob("*"))
        if p.is_file()
    }


def _prepare_old_dist(out: Path) -> dict[str, bytes]:
    (out / "livres").mkdir(parents=True)
    (out / "assets").mkdir()
    (out / "covers").mkdir()
    (out / "onix").mkdir()
    (out / "livres" / "orpheline.html").write_text("ancienne page", encoding="utf-8")
    (out / "index.html").write_text("ancien index", encoding="utf-8")
    (out / "manuel.txt").write_text("manuel", encoding="utf-8")
    (out / "assets" / "manuel-asset.txt").write_text("asset manuel", encoding="utf-8")
    (out / "covers" / "cover.jpg").write_bytes(b"cover")
    (out / "onix" / "temoin.xml").write_text("<ONIX/>", encoding="utf-8")
    return _snapshot(out)


def test_pages_editoriales_renommees_et_supprimees_disparaissent(tmp_path):
    out = tmp_path / "site-sortie"
    wb1 = _workbook(
        tmp_path / "site-1.xlsx",
        pages_rows=[
            {"slug": "presentation", "title": "Presentation", "content_md": "Texte"},
            {"slug": "ancienne-page", "title": "Ancienne", "content_md": "Ancienne"},
        ],
    )
    bs.build_site(wb1, out, covers_dir=None)
    assert (out / "presentation.html").exists()
    assert (out / "ancienne-page.html").exists()

    wb2 = _workbook(
        tmp_path / "site-2.xlsx",
        pages_rows=[
            {"slug": "la-maison", "title": "La maison", "content_md": "Texte"},
        ],
    )
    bs.build_site(wb2, out, covers_dir=None)

    assert (out / "la-maison.html").exists()
    assert not (out / "presentation.html").exists()
    assert not (out / "ancienne-page.html").exists()


def test_nettoyage_complet_supprime_tout_sauf_assets_et_covers(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "site-sortie"
    for rel, content in {
        "index.html": "ancien",
        "ancienne-page.html": "ancien",
        "ancien-fichier.json": "{}",
        "validation.csv": "ancien",
        "livres/ancien.html": "ancien",
        "collections/ancienne.html": "ancien",
        "revues/ancienne.html": "ancien",
        "onix/ancien.xml": "<onix/>",
        "manuel-racine.pdf": "PDF",
        "assets/logo.png": "PNG",
        "assets/docs/document.pdf": "PDF",
        "assets/catalogue.json": "{}",
        "assets/actualites.json": "[]",
        "assets/sous-dossier/fichier.txt": "asset",
        "covers/couverture.jpg": "cover",
    }.items():
        p = out / rel
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(content, encoding="utf-8")

    bs.build_site(wb, out, covers_dir=None)

    for rel in [
        "ancienne-page.html", "ancien-fichier.json", "livres/ancien.html",
        "collections/ancienne.html", "revues/ancienne.html", "onix/ancien.xml",
        "manuel-racine.pdf", "assets/catalogue.json", "assets/actualites.json",
    ]:
        assert not (out / rel).exists()

    for rel in [
        "assets/logo.png", "assets/docs/document.pdf",
        "assets/sous-dossier/fichier.txt", "covers/couverture.jpg",
    ]:
        assert (out / rel).exists()

    for rel in [
        "index.html", "catalogue.html", "catalogue.json", "actualites.json",
        "validation.csv", "livres", "collections", "revues",
    ]:
        assert (out / rel).exists()


def test_dossier_sortie_nom_arbitraire(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "site-purh-public"

    bs.build_site(wb, out, covers_dir=None)

    assert (out / "index.html").exists()
    assert (out / "catalogue.json").exists()
    assert _leftovers(out) == []


def test_assets_ou_covers_non_repertoire_bloquent_sans_modifier_sortie(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx")
    for name in ("assets", "covers"):
        out = tmp_path / f"sortie-{name}"
        out.mkdir()
        (out / name).write_text("pas un dossier", encoding="utf-8")
        before = _snapshot(out)

        with pytest.raises(RuntimeError, match="pas un dossier"):
            bs.build_site(wb, out, covers_dir=None)

        assert _snapshot(out) == before
        assert _leftovers(out) == []


def test_transaction_reussie_sans_ancien_dossier(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "site-sortie"

    report = bs.build_site(wb, out, covers_dir=None)

    assert out.exists()
    assert (out / "index.html").exists()
    assert (out / "validation.csv").exists()
    assert report is not None
    assert _leftovers(out) == []


def test_transaction_reussie_recompose_sortie_et_conserve_assets_covers(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "site-sortie"
    _prepare_old_dist(out)
    (out / "assets" / "catalogue.json").write_text("ancien catalogue", encoding="utf-8")
    (out / "assets" / "actualites.json").write_text("anciennes actus", encoding="utf-8")

    bs.build_site(wb, out, covers_dir=None)

    assert (out / "index.html").read_text(encoding="utf-8") != "ancien index"
    assert not (out / "livres" / "orpheline.html").exists()
    assert not (out / "manuel.txt").exists()
    assert (out / "assets" / "manuel-asset.txt").read_text(encoding="utf-8") == "asset manuel"
    assert not (out / "assets" / "catalogue.json").exists()
    assert not (out / "assets" / "actualites.json").exists()
    assert (out / "covers" / "cover.jpg").read_bytes() == b"cover"
    assert not (out / "onix" / "temoin.xml").exists()
    assert (out / "catalogue.json").exists()
    assert (out / "actualites.json").exists()
    assert _leftovers(out) == []


def test_echec_generation_garde_ancien_dist_intact_et_pas_de_ftp(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "dist"
    before = _prepare_old_dist(out)
    ftp_calls = []

    def fail_catalogue(*args, **kwargs):
        raise RuntimeError("boom-generation")

    monkeypatch.setattr(orchestrator, "build_catalogue_page", fail_catalogue)
    monkeypatch.setattr(orchestrator, "publish_ftp", lambda *a, **k: ftp_calls.append((a, k)))

    with pytest.raises(RuntimeError, match="boom-generation"):
        bs.build_site(wb, out, covers_dir=None, publish=True)

    assert _snapshot(out) == before
    assert _leftovers(out) == []
    assert ftp_calls == []


def test_echec_basculement_restaure_ancien_dist(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "dist"
    before = _prepare_old_dist(out)
    original_rename = output_transaction._rename_path

    def flaky_rename(src: Path, dst: Path) -> None:
        if ".build-" in src.name and dst == out:
            raise RuntimeError("boom-rename")
        original_rename(src, dst)

    monkeypatch.setattr(output_transaction, "_rename_path", flaky_rename)

    with pytest.raises(RuntimeError, match="boom-rename"):
        bs.build_site(wb, out, covers_dir=None)

    assert _snapshot(out) == before
    assert _leftovers(out) == []


def test_permissionerror_basculement_restaure_sans_copie_de_secours(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "dist"
    before = _prepare_old_dist(out)
    original_rename = output_transaction._rename_path
    original_copytree = output_transaction.shutil.copytree
    copy_calls = []

    def observed_copytree(src: Path, dst: Path, *args, **kwargs):
        copy_calls.append((Path(src), Path(dst)))
        return original_copytree(src, dst, *args, **kwargs)

    def flaky_rename(src: Path, dst: Path) -> None:
        if ".build-" in src.name and dst == out:
            raise PermissionError("boom-permission")
        original_rename(src, dst)

    monkeypatch.setattr(output_transaction.shutil, "copytree", observed_copytree)
    monkeypatch.setattr(output_transaction, "_rename_path", flaky_rename)

    with pytest.raises(PermissionError, match="boom-permission"):
        bs.build_site(wb, out, covers_dir=None)

    assert _snapshot(out) == before
    assert _leftovers(out) == []
    assert all(dst != out for _src, dst in copy_calls)


def test_permissionerror_sans_ancien_out_ne_cree_pas_de_sortie_partielle(tmp_path, monkeypatch):
    out = tmp_path / "dist"
    original_rename = output_transaction._rename_path

    def flaky_rename(src: Path, dst: Path) -> None:
        if ".build-" in src.name and dst == out:
            raise PermissionError("boom-permission")
        original_rename(src, dst)

    monkeypatch.setattr(output_transaction, "_rename_path", flaky_rename)

    with pytest.raises(PermissionError, match="boom-permission"):
        with output_transaction.staged_output(out) as tx:
            (tx.staging_dir / "index.html").write_text("nouveau", encoding="utf-8")
            tx.commit()

    assert not out.exists()
    assert _leftovers(out) == []


def test_echec_nettoyage_backup_signale_et_garde_nouveau_site(tmp_path, monkeypatch, recwarn):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "dist"
    _prepare_old_dist(out)
    original_remove = output_transaction._remove_tree

    def flaky_remove(path: Path) -> None:
        if ".backup-" in path.name:
            raise RuntimeError("boom-cleanup")
        original_remove(path)

    monkeypatch.setattr(output_transaction, "_remove_tree", flaky_remove)

    bs.build_site(wb, out, covers_dir=None)

    assert (out / "index.html").exists()
    assert "ancien index" not in (out / "index.html").read_text(encoding="utf-8")
    warnings = [w for w in recwarn if issubclass(w.category, output_transaction.OutputBackupCleanupWarning)]
    assert warnings
    backups = sorted(out.parent.glob(f".{out.name}.backup-*"))
    assert len(backups) == 1
    original_remove(backups[0])


def test_warning_backup_en_erreur_ne_declenche_pas_rollback(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "dist"
    _prepare_old_dist(out)
    original_remove = output_transaction._remove_tree

    def flaky_remove(path: Path) -> None:
        if ".backup-" in path.name:
            raise RuntimeError("boom-cleanup")
        original_remove(path)

    monkeypatch.setattr(output_transaction, "_remove_tree", flaky_remove)

    with warnings.catch_warnings():
        warnings.simplefilter("error", output_transaction.OutputBackupCleanupWarning)
        with pytest.raises(output_transaction.OutputBackupCleanupWarning):
            bs.build_site(wb, out, covers_dir=None)

    assert (out / "index.html").exists()
    assert "ancien index" not in (out / "index.html").read_text(encoding="utf-8")
    assert not sorted(out.parent.glob(f".{out.name}.build-*"))
    backups = sorted(out.parent.glob(f".{out.name}.backup-*"))
    assert len(backups) == 1
    assert (backups[0] / "index.html").read_text(encoding="utf-8") == "ancien index"
    original_remove(backups[0])


def test_echec_preparation_staging_nettoie_copie_partielle(tmp_path, monkeypatch):
    out = tmp_path / "dist"
    before = _prepare_old_dist(out)
    original_copytree = output_transaction.shutil.copytree

    def failing_copytree(src: Path, dst: Path, *args, **kwargs):
        dst = Path(dst)
        dst.mkdir(parents=True)
        (dst / "partiel.txt").write_text("partiel", encoding="utf-8")
        raise RuntimeError("boom-copytree")

    monkeypatch.setattr(output_transaction.shutil, "copytree", failing_copytree)

    with pytest.raises(RuntimeError, match="boom-copytree"):
        with output_transaction.staged_output(out):
            pass

    assert _snapshot(out) == before
    assert _leftovers(out) == []
    monkeypatch.setattr(output_transaction.shutil, "copytree", original_copytree)


def test_echec_nettoyage_preserve_erreur_initiale_dans_message(tmp_path, monkeypatch):
    out = tmp_path / "dist"
    original_remove = output_transaction._remove_tree
    staging_seen = {}

    def flaky_remove(path: Path) -> None:
        if ".build-" in path.name:
            staging_seen["path"] = path
            raise RuntimeError("boom-rmtree")
        original_remove(path)

    monkeypatch.setattr(output_transaction, "_remove_tree", flaky_remove)

    with pytest.raises(output_transaction.OutputCleanupError) as exc:
        with output_transaction.staged_output(out) as tx:
            (tx.staging_dir / "partiel.txt").write_text("partiel", encoding="utf-8")
            raise RuntimeError("boom-initial")

    msg = str(exc.value)
    assert "boom-initial" in msg
    assert "Staging résiduel" in msg
    assert str(staging_seen["path"]) in msg
    original_remove(staging_seen["path"])


def test_validate_only_ne_laisse_pas_de_staging_global(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx", title="")
    out = tmp_path / "sortie"
    (out / "assets").mkdir(parents=True)
    (out / "assets" / "catalogue.json").write_text("ancien", encoding="utf-8")
    (out / "ancien-site.html").write_text("ancien", encoding="utf-8")
    (out / "onix").mkdir()
    (out / "onix" / "ancien.xml").write_text("<onix/>", encoding="utf-8")

    bs.build_site(wb, out, covers_dir=None, validate_only=True, force_alerts=True)

    assert (out / "validation.csv").exists()
    assert (out / "catalogue.json").exists()
    assert not (out / "assets" / "catalogue.json").exists()
    assert (out / "ancien-site.html").exists()
    assert (out / "onix" / "ancien.xml").exists()
    assert not (out / "index.html").exists()
    assert _leftovers(out) == []


def test_alerte_sans_force_ne_remplace_pas_ancien_site(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx", title="")
    out = tmp_path / "dist"
    before = _prepare_old_dist(out)

    with pytest.raises(bs.ValidationAlertError):
        bs.build_site(wb, out, covers_dir=None, force_alerts=False)

    after = _snapshot(out)
    assert after["index.html"] == before["index.html"]
    assert after["livres/orpheline.html"] == before["livres/orpheline.html"]
    assert "validation.csv" in after
    assert _leftovers(out) == []


def test_alerte_avec_force_genere_transactionnellement(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx", title="")
    out = tmp_path / "dist"
    _prepare_old_dist(out)

    report = bs.build_site(wb, out, covers_dir=None, force_alerts=True)

    assert report.has_alerts
    assert (out / "index.html").exists()
    assert not (out / "livres" / "orpheline.html").exists()
    assert _leftovers(out) == []
