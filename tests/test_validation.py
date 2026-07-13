import sys
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import build_site as bs
import cidre.orchestrator as orchestrator
from cidre.data_models import load_config
from cidre.excel_data import detect_books_sheet, load_books
from gui_tk import (
    ASSETS_HELP_TEXT,
    App,
    DEFAULT_ASSETS_VALUE,
    find_book_url_collisions,
    format_blocking_validation_message,
    resolve_assets_dir_for_gui,
    should_continue_after_validation,
)
import gui_tk


def _book(**over):
    row = {
        "slug": "un-livre",
        "id13": "9782877750001",
        "titre_norm": "Un livre",
        "sous_titre_norm": "",
        "credit_ligne": "",
        "collection": "Essais",
        "collection_id": "col-essais",
        "date_parution_norm": "2026-01-01",
        "pub_date": pd.Timestamp("2026-01-01").date(),
        "format_site": "",
        "price": "12",
        "prix_ttc": "",
        "availability": "",
        "availability_label": "",
        "cover_file": "cover.jpg",
        "Description courte": "Resume",
        "Description longue": "",
        "Table des matieres": "",
        "order_url": "",
        "openedition_url": "",
    }
    row.update(over)
    return row


def _report(books, **kwargs):
    return bs.validate_site_data(books=pd.DataFrame(books), **kwargs)


def _codes(report):
    return {i.code for i in report.issues}


def _loaded_books(path: Path) -> pd.DataFrame:
    with pd.ExcelFile(path) as wb:
        cfg = load_config(wb, "CONFIG")
        return load_books(wb, detect_books_sheet(wb, cfg.books_sheet))


def _leftovers(out: Path) -> list[Path]:
    return sorted(out.parent.glob(f".{out.name}.build-*")) + sorted(out.parent.glob(f".{out.name}.backup-*"))


def _snapshot(path: Path) -> dict[str, bytes]:
    return {
        p.relative_to(path).as_posix(): p.read_bytes()
        for p in sorted(path.rglob("*"))
        if p.is_file()
    }


def test_donnees_valides_sans_blocage_ni_alerte(tmp_path):
    covers = tmp_path / "covers"
    covers.mkdir()
    (covers / "cover.jpg").write_bytes(b"jpg")
    collections = pd.DataFrame([{"collection_id": "col-essais", "is_active": 1, "slug": "essais"}])
    report = _report([_book()], collections=collections, covers_dir=covers)
    assert not report.has_blocking_issues
    assert not report.has_alerts


def test_titre_absent_alerte_contournable():
    report = _report([_book(titre_norm="")])
    assert "BOOK_TITLE_MISSING" in _codes(report)
    assert report.alerts[0].severity == "alert"


def test_isbn_absent_avertissement():
    report = _report([_book(id13="")])
    assert "BOOK_ID13_MISSING" in _codes(report)
    assert not report.has_alerts


def test_resume_absent_avertissement():
    report = _report([_book(**{"Description courte": "", "Description longue": ""})])
    assert "BOOK_DESCRIPTION_MISSING" in _codes(report)
    assert not report.has_alerts


def test_couverture_absente_ou_introuvable(tmp_path):
    assert "BOOK_COVER_MISSING" in _codes(_report([_book(cover_file="")]))
    covers = tmp_path / "covers"
    covers.mkdir()
    report = _report([_book(cover_file="absente.jpg")], covers_dir=covers)
    assert "BOOK_COVER_NOT_FOUND" in _codes(report)


def test_slug_duplique_alerte():
    report = _report([_book(slug="x"), _book(slug="x", id13="9782877750002")])
    assert "BOOK_SLUG_DUPLICATE" in _codes(report)
    assert report.has_alerts


def test_isbn_duplique_alerte():
    report = _report([_book(slug="a"), _book(slug="b")])
    assert "BOOK_ID13_DUPLICATE" in _codes(report)
    assert report.has_alerts


def test_collection_inconnue_alerte():
    collections = pd.DataFrame([{"collection_id": "col-connue", "is_active": 1, "slug": "connue"}])
    report = _report([_book(collection_id="col-inconnue")], collections=collections)
    assert "BOOK_COLLECTION_UNKNOWN" in _codes(report)


def test_revue_inconnue_alerte():
    revues = pd.DataFrame([{"journal_id": "rev-connue", "is_active": 1, "slug": "connue"}])
    report = _report([
        _book(collection="Revue inconnue", collection_id="rev-inconnue", titre_norm="Revue inconnue n° 1")
    ], revues=revues)
    assert "BOOK_REVUE_UNKNOWN" in _codes(report)


def test_date_invalide_avertissement():
    report = _report([_book(date_parution_norm="pas-une-date")])
    assert "BOOK_DATE_INVALID" in _codes(report)


def test_url_invalide_avertissement():
    report = _report([_book(order_url="javascript:alert(1)")])
    assert "BOOK_URL_INVALID" in _codes(report)


def test_generation_autorisee_avec_seulement_avertissements(tmp_path):
    wb = _workbook(tmp_path / "warn.xlsx", title="Un livre", id13="")
    out = tmp_path / "dist"
    bs.build_site(wb, out, covers_dir=None, force_alerts=False)
    assert (out / "index.html").exists()
    csv = (out / "validation.csv").read_text(encoding="utf-8")
    assert "BOOK_ID13_MISSING" in csv


def test_cli_refuse_alerte_sans_force(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "alert.xlsx", title="")
    out = tmp_path / "dist"
    monkeypatch.setattr(sys, "argv", ["build_site.py", "--excel", str(wb), "--out", str(out)])
    with pytest.raises(SystemExit) as exc:
        bs.main()
    assert exc.value.code == 4
    assert (out / "validation.csv").exists()
    assert not (out / "index.html").exists()


def test_cli_out_fichier_existant_blocage_sans_modifier(tmp_path, monkeypatch, capsys):
    wb = _workbook(tmp_path / "ok.xlsx")
    out = tmp_path / "dist.html"
    out.write_text("ancien", encoding="utf-8")
    monkeypatch.setattr(sys, "argv", ["build_site.py", "--excel", str(wb), "--out", str(out)])
    with pytest.raises(SystemExit) as exc:
        bs.main()
    assert exc.value.code == 3
    assert out.read_text(encoding="utf-8") == "ancien"
    err = capsys.readouterr().err
    assert "OUTPUT_PATH_NOT_DIRECTORY" in err
    assert "Traceback" not in err


def test_cli_out_ancetre_fichier_blocage_sans_creation(tmp_path, monkeypatch, capsys):
    wb = _workbook(tmp_path / "ok.xlsx")
    parent_file = tmp_path / "fichier-existant.txt"
    parent_file.write_text("ancien", encoding="utf-8")
    out = parent_file / "sous-dossier"
    monkeypatch.setattr(sys, "argv", ["build_site.py", "--excel", str(wb), "--out", str(out)])
    with pytest.raises(SystemExit) as exc:
        bs.main()
    assert exc.value.code == 3
    assert parent_file.read_text(encoding="utf-8") == "ancien"
    assert not out.exists()
    err = capsys.readouterr().err
    assert "OUTPUT_PARENT_NOT_DIRECTORY" in err
    assert "Traceback" not in err


def test_cli_autorise_alerte_avec_force(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "alert-force.xlsx", title="")
    out = tmp_path / "dist"
    monkeypatch.setattr(sys, "argv", [
        "build_site.py", "--excel", str(wb), "--out", str(out), "--force"
    ])
    bs.main()
    assert (out / "index.html").exists()
    assert "BOOK_TITLE_MISSING" in (out / "validation.csv").read_text(encoding="utf-8")


def test_cli_assets_dir_absent_autorise_generation(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "ok.xlsx")
    out = tmp_path / "site-sortie"
    monkeypatch.setattr(sys, "argv", ["build_site.py", "--excel", str(wb), "--out", str(out)])
    bs.main()
    assert (out / "index.html").exists()


def test_cli_assets_dir_valide_copie_recursivement(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "ok.xlsx")
    out = tmp_path / "site-sortie"
    assets = tmp_path / "assets-source"
    (assets / "docs").mkdir(parents=True)
    (assets / "docs" / "consignes.pdf").write_text("pdf", encoding="utf-8")
    (assets / "logo.png").write_text("logo", encoding="utf-8")
    monkeypatch.setattr(sys, "argv", [
        "build_site.py", "--excel", str(wb), "--out", str(out), "--assets-dir", str(assets)
    ])
    bs.main()
    assert (out / "assets" / "docs" / "consignes.pdf").read_text(encoding="utf-8") == "pdf"
    assert (out / "assets" / "logo.png").read_text(encoding="utf-8") == "logo"


def test_cli_assets_dir_inexistant_ou_fichier_refuse(tmp_path, monkeypatch, capsys):
    wb = _workbook(tmp_path / "ok.xlsx")
    out = tmp_path / "site-sortie"
    for assets in [tmp_path / "absent", tmp_path / "source.txt"]:
        if assets.suffix:
            assets.write_text("pas un dossier", encoding="utf-8")
        monkeypatch.setattr(sys, "argv", [
            "build_site.py", "--excel", str(wb), "--out", str(out), "--assets-dir", str(assets)
        ])
        with pytest.raises(SystemExit) as exc:
            bs.main()
        assert exc.value.code == 3
        assert "Dossier des assets invalide" in capsys.readouterr().err
        assert not (out / "index.html").exists()


def test_cli_assets_dir_imbrique_refuse(tmp_path, monkeypatch, capsys):
    wb = _workbook(tmp_path / "ok.xlsx")
    out = tmp_path / "site-sortie"
    cases = [
        out,
        out / "assets",
        out / "sous-dossier",
        tmp_path,
    ]
    for assets in cases:
        assets.mkdir(parents=True, exist_ok=True)
        monkeypatch.setattr(sys, "argv", [
            "build_site.py", "--excel", str(wb), "--out", str(out), "--assets-dir", str(assets)
        ])
        with pytest.raises(SystemExit) as exc:
            bs.main()
        assert exc.value.code == 3
        assert "Dossier des assets invalide" in capsys.readouterr().err
        assert not (out / "index.html").exists()


def test_blocage_technique_empeche_ecriture_ancien_out(tmp_path):
    wb = _workbook(tmp_path / "ok.xlsx")
    out = tmp_path / "dist"
    out.write_text("ancien", encoding="utf-8")
    with pytest.raises(bs.ValidationBlockingError):
        bs.build_site(wb, out, covers_dir=None)
    assert out.read_text(encoding="utf-8") == "ancien"


def test_pas_de_ftp_en_cas_alerte_cli(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "ftp-alert.xlsx", title="", ftp=True)
    out = tmp_path / "dist"
    calls = []
    monkeypatch.setattr(orchestrator, "publish_ftp", lambda *a, **k: calls.append((a, k)))
    monkeypatch.setattr(sys, "argv", [
        "build_site.py", "--excel", str(wb), "--out", str(out), "--publish-ftp"
    ])
    with pytest.raises(SystemExit) as exc:
        bs.main()
    assert exc.value.code == 4
    assert calls == []


def test_validate_only_meme_moteur_sans_html(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "validate-only.xlsx", title="")
    out = tmp_path / "dist"
    monkeypatch.setattr(sys, "argv", [
        "build_site.py", "--excel", str(wb), "--out", str(out), "--validate-only", "--force"
    ])
    bs.main()
    assert "BOOK_TITLE_MISSING" in (out / "validation.csv").read_text(encoding="utf-8")
    assert (out / "catalogue.json").exists()
    assert not (out / "assets" / "catalogue.json").exists()
    assert not (out / "index.html").exists()


def test_slug_duplique_depuis_classeur_apres_unicisation(tmp_path, monkeypatch, capsys):
    wb = _workbook(
        tmp_path / "dup-slug.xlsx",
        book_rows=[
            {"slug": "meme-slug", "id13": "9782877750001", "titre_norm": "Livre A"},
            {"slug": "meme-slug", "id13": "9782877750002", "titre_norm": "Livre B"},
        ],
    )
    out = tmp_path / "dist"
    monkeypatch.setattr(sys, "argv", ["build_site.py", "--excel", str(wb), "--out", str(out)])
    with pytest.raises(SystemExit) as exc:
        bs.main()
    assert exc.value.code == 3
    err = capsys.readouterr().err
    assert "BOOK_SLUG_DUPLICATE" in err or "1 alerte" in err
    assert "DUPLICATE_OUTPUT_TARGET" in err
    assert not (out / "index.html").exists()


def test_livres_meme_slug_explicite_conservent_le_meme_slug_et_bloquent(tmp_path):
    wb = _workbook(
        tmp_path / "dup-books.xlsx",
        book_rows=[
            {"slug": "meme-slug", "id13": "9782877750001", "titre_norm": "Livre A"},
            {"slug": "meme-slug", "id13": "9782877750002", "titre_norm": "Livre B"},
        ],
    )
    books = _loaded_books(wb)
    assert books["_slug_candidate"].tolist() == ["meme-slug", "meme-slug"]
    assert books["slug"].tolist() == ["meme-slug", "meme-slug"]
    assert books["_slug_was_uniquified"].tolist() == [False, False]
    report = bs.validate_site_data(books=books)
    assert "DUPLICATE_OUTPUT_TARGET" in _codes(report)
    assert any(i.severity == "blocking" for i in report.issues if i.code == "DUPLICATE_OUTPUT_TARGET")


def test_reordonnancement_livres_doublons_reste_bloquant_sans_suffixe(tmp_path):
    rows_ab = [
        {"slug": "meme-slug", "id13": "9782877750001", "titre_norm": "Livre A"},
        {"slug": "meme-slug", "id13": "9782877750002", "titre_norm": "Livre B"},
    ]
    rows_ba = list(reversed(rows_ab))
    for name, rows in [("ab.xlsx", rows_ab), ("ba.xlsx", rows_ba)]:
        wb = _workbook(tmp_path / name, book_rows=rows)
        books = _loaded_books(wb)
        assert books["_slug_candidate"].tolist() == ["meme-slug", "meme-slug"]
        assert books["slug"].tolist() == ["meme-slug", "meme-slug"]
        with pytest.raises(bs.ValidationBlockingError):
            bs.build_site(wb, tmp_path / f"dist-{name}", covers_dir=None, force_alerts=True)


def test_livre_inactif_meme_slug_ne_cree_pas_de_collision(tmp_path):
    wb = _workbook(
        tmp_path / "inactive.xlsx",
        book_rows=[
            {"slug": "meme-slug", "id13": "9782877750001", "titre_norm": "Livre actif", "active_site": 1},
            {"slug": "meme-slug", "id13": "9782877750002", "titre_norm": "Livre inactif", "active_site": 0},
        ],
    )
    books = _loaded_books(wb)
    assert books["titre_norm"].tolist() == ["Livre actif"]
    report = bs.validate_site_data(books=books)
    assert "DUPLICATE_OUTPUT_TARGET" not in _codes(report)


def test_fallbacks_livres_conserves_et_collision_bloquante(tmp_path):
    wb = _workbook(
        tmp_path / "fallbacks.xlsx",
        book_rows=[
            {"slug": "", "id13": "9782877750001", "titre_norm": "Titre commun"},
            {"slug": "", "id13": "9782877750002", "titre_norm": "Titre commun"},
            {"slug": "", "id13": "", "titre_norm": "Sans ISBN"},
            {"slug": "", "id13": "", "titre_norm": "Sans ISBN"},
        ],
    )
    books = _loaded_books(wb)
    assert books["slug"].tolist() == [
        "titre-commun-9782877750001",
        "titre-commun-9782877750002",
        "sans-isbn",
        "sans-isbn",
    ]
    assert books["_slug_candidate"].tolist() == books["slug"].tolist()
    assert books["_slug_was_uniquified"].tolist() == [False, False, False, False]
    report = bs.validate_site_data(books=books)
    assert "DUPLICATE_OUTPUT_TARGET" in _codes(report)


def test_validation_collision_livres_bloque_transaction_et_ftp(tmp_path, monkeypatch):
    wb = _workbook(
        tmp_path / "dup-tx.xlsx",
        book_rows=[
            {"slug": "meme-slug", "id13": "9782877750001", "titre_norm": "Livre A"},
            {"slug": "meme-slug", "id13": "9782877750002", "titre_norm": "Livre B"},
        ],
    )
    out = tmp_path / "dist"
    (out / "livres").mkdir(parents=True)
    (out / "livres" / "ancien.html").write_text("ancien", encoding="utf-8")
    before = _snapshot(out)
    ftp_calls = []
    monkeypatch.setattr(orchestrator, "publish_ftp", lambda *a, **k: ftp_calls.append((a, k)))

    with pytest.raises(bs.ValidationBlockingError):
        bs.build_site(wb, out, covers_dir=None, force_alerts=True, publish=True)

    assert _snapshot(out) == before
    assert ftp_calls == []
    assert _leftovers(out) == []


def test_slug_genere_avertissement_sans_blocage(tmp_path):
    wb = _workbook(
        tmp_path / "generated-slug.xlsx",
        book_rows=[
            {"slug": "", "id13": "9782877750001", "titre_norm": "Titre sans slug"},
        ],
    )
    books = _loaded_books(wb)
    report = bs.validate_site_data(books=books)
    issues = [i for i in report.issues if i.code == "BOOK_SLUG_GENERATED"]
    assert len(issues) == 1
    assert issues[0].severity == "warning"
    assert not report.has_blocking_issues
    out = tmp_path / "dist"
    bs.build_site(wb, out, covers_dir=None, force_alerts=False)
    assert (out / "livres" / "titre-sans-slug-9782877750001.html").exists()


def test_pages_meme_slug_et_identifiant_collision_blocage():
    pages = pd.DataFrame([
        {"slug": "doublon", "title": "Doublon"},
        {"slug": "doublon", "title": "Doublon"},
    ])
    report = bs.validate_site_data(books=pd.DataFrame(), pages=pages)
    assert "DUPLICATE_OUTPUT_TARGET" in _codes(report)
    duplicate_issues = [i for i in report.blocking_issues if i.code == "DUPLICATE_OUTPUT_TARGET"]
    assert len(duplicate_issues) == 1
    issue = duplicate_issues[0]
    assert "page:doublon" in issue.message


def test_page_non_publiee_catalogue_ne_bloque_pas():
    pages = pd.DataFrame([{"slug": "catalogue", "title": "Catalogue manuel", "is_published": 0}])
    report = bs.validate_site_data(books=pd.DataFrame(), pages=pages)
    assert not report.has_blocking_issues


def test_deux_pages_non_publiees_meme_slug_ne_bloquent_pas():
    pages = pd.DataFrame([
        {"slug": "doublon", "title": "Doublon", "is_published": 0},
        {"slug": "doublon", "title": "Doublon", "is_published": 0},
    ])
    report = bs.validate_site_data(books=pd.DataFrame(), pages=pages)
    assert not report.has_blocking_issues


def test_page_publiee_catalogue_bloque_toujours():
    pages = pd.DataFrame([{"slug": "catalogue", "title": "Catalogue manuel", "is_published": 1}])
    report = bs.validate_site_data(books=pd.DataFrame(), pages=pages)
    assert "DUPLICATE_OUTPUT_TARGET" in _codes(report)
    assert any("catalogue.html" in i.message for i in report.blocking_issues)


def test_deux_pages_publiees_meme_slug_bloquent_toujours():
    pages = pd.DataFrame([
        {"slug": "doublon", "title": "Doublon", "is_published": 1},
        {"slug": "doublon", "title": "Doublon", "is_published": 1},
    ])
    report = bs.validate_site_data(books=pd.DataFrame(), pages=pages)
    duplicate_issues = [i for i in report.blocking_issues if i.code == "DUPLICATE_OUTPUT_TARGET"]
    assert len(duplicate_issues) == 1
    assert "doublon.html" in duplicate_issues[0].message


def test_collections_meme_slug_et_identifiant_collision_blocage():
    collections = pd.DataFrame([
        {"collection_id": "col", "slug": "index", "is_active": 1},
        {"collection_id": "col", "slug": "index", "is_active": 1},
    ])
    report = bs.validate_site_data(books=pd.DataFrame(), collections=collections)
    assert "DUPLICATE_OUTPUT_TARGET" in _codes(report)
    duplicate_issues = [i for i in report.blocking_issues if i.code == "DUPLICATE_OUTPUT_TARGET"]
    assert len(duplicate_issues) == 1
    assert "collections/index.html" in duplicate_issues[0].message


def test_revues_meme_slug_et_identifiant_collision_blocage():
    revues = pd.DataFrame([
        {"journal_id": "rev", "slug": "index", "title": "Revue", "is_active": 1},
        {"journal_id": "rev", "slug": "index", "title": "Revue", "is_active": 1},
    ])
    report = bs.validate_site_data(books=pd.DataFrame(), revues=revues)
    assert "DUPLICATE_OUTPUT_TARGET" in _codes(report)
    duplicate_issues = [i for i in report.blocking_issues if i.code == "DUPLICATE_OUTPUT_TARGET"]
    assert len(duplicate_issues) == 1
    assert "revues/index.html" in duplicate_issues[0].message


def test_page_collision_avec_page_automatique_blocage():
    pages = pd.DataFrame([{"slug": "catalogue", "title": "Catalogue manuel"}])
    report = bs.validate_site_data(books=pd.DataFrame(), pages=pages)
    assert "DUPLICATE_OUTPUT_TARGET" in _codes(report)
    assert any("catalogue.html" in i.message for i in report.blocking_issues)


def test_page_pilotee_par_pages_autorisee_si_producteur_unique():
    pages = pd.DataFrame([
        {"slug": "presentation", "title": "Presentation"},
        {"slug": "commander", "title": "Commander"},
        {"slug": "open-access", "title": "Open Access"},
    ])
    report = bs.validate_site_data(books=pd.DataFrame(), pages=pages)
    assert not report.has_blocking_issues


def test_page_contact_publiee_autorisee_et_generable(tmp_path):
    wb = _workbook(
        tmp_path / "contact.xlsx",
        pages_rows=[
            {"slug": "presentation", "title": "Presentation", "content_md": "Texte", "is_published": 1},
            {"slug": "contact", "title": "Contact", "content_md": "Nous contacter", "is_published": 1},
        ],
    )
    out = tmp_path / "dist"
    bs.build_site(wb, out, covers_dir=None)
    assert (out / "contact.html").exists()


def test_csv_colonnes_attendues(tmp_path):
    report = _report([_book(id13="")])
    p = tmp_path / "validation.csv"
    bs.write_validation_csv(report, p)
    header = p.read_text(encoding="utf-8").splitlines()[0]
    assert header == "level,code,entity,identifier,field,message"


def test_imports_historiques_depuis_build_site():
    assert bs.build_validation_report
    assert bs.ValidationIssue
    assert bs.ValidationReport
    assert bs.validate_site_data


def test_decision_gui_pure():
    report = _report([_book(titre_norm="")])
    assert should_continue_after_validation(report, lambda r: True) is True
    assert should_continue_after_validation(report, lambda r: False) is False
    blocking = bs.ValidationReport([
        bs.ValidationIssue("blocking", "X", "site", "id", "field", "msg")
    ])
    assert should_continue_after_validation(blocking, lambda r: True) is False


def test_gui_assets_champ_vide_par_defaut_et_aide():
    assert DEFAULT_ASSETS_VALUE == ""
    for expected in ["docs/", "images/", "actu/", "social/", "second dossier nommé « assets »", "couvertures", "catalogue.json", "actualites.json"]:
        assert expected in ASSETS_HELP_TEXT


def test_gui_assets_resolve_facultatif_valide_et_invalide(tmp_path):
    out = tmp_path / "site-sortie"
    assets = tmp_path / "assets-source"
    assets.mkdir()
    assert resolve_assets_dir_for_gui("", out) is None
    assert resolve_assets_dir_for_gui(str(assets), out) == assets.resolve()
    with pytest.raises(bs.AssetSourceError):
        resolve_assets_dir_for_gui(str(tmp_path / "absent"), out)


class _FakeVar:
    def __init__(self, value):
        self.value = value

    def get(self):
        return self.value

    def set(self, value):
        self.value = value


class _FakeButton:
    def config(self, **kwargs):
        pass


class _ImmediateThread:
    def __init__(self, target, daemon=None):
        self.target = target

    def start(self):
        self.target()


def test_gui_run_build_transmet_assets_dir(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "site-sortie"
    assets = tmp_path / "assets-source"
    assets.mkdir()
    captured = {}

    app = object.__new__(App)
    app.var_excel = _FakeVar(str(wb))
    app.var_out = _FakeVar(str(out))
    app.var_covers = _FakeVar("")
    app.var_assets = _FakeVar(str(assets))
    app.var_validate_only = _FakeVar(False)
    app.var_publish_ftp = _FakeVar(False)
    app.var_export_onix = _FakeVar(False)
    app.var_start_server = _FakeVar(False)
    app.var_port = _FakeVar(8000)
    app.btn_run = _FakeButton()
    app._preview_server = None
    app.log = lambda msg: None
    app.after = lambda delay, func=None: func() if func else None
    app.stop_server = lambda: None

    def fake_build_site(**kwargs):
        captured.update(kwargs)
        return None

    monkeypatch.setattr(gui_tk, "build_site", fake_build_site)
    monkeypatch.setattr(gui_tk.threading, "Thread", _ImmediateThread)

    app.run_build()

    assert captured["assets_dir"] == assets.resolve()
    assert captured["out_dir"] == out.resolve()


def test_gui_run_build_assets_invalide_arrete_generation(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx")
    out = tmp_path / "site-sortie"
    errors = []
    called = []

    app = object.__new__(App)
    app.var_excel = _FakeVar(str(wb))
    app.var_out = _FakeVar(str(out))
    app.var_covers = _FakeVar("")
    app.var_assets = _FakeVar(str(tmp_path / "absent"))
    app.var_validate_only = _FakeVar(False)
    app.var_publish_ftp = _FakeVar(False)
    app.var_export_onix = _FakeVar(False)
    app.var_start_server = _FakeVar(False)
    app.var_port = _FakeVar(8000)

    monkeypatch.setattr(gui_tk.messagebox, "showerror", lambda title, message: errors.append((title, message)))
    monkeypatch.setattr(gui_tk, "build_site", lambda **kwargs: called.append(kwargs))

    app.run_build()

    assert errors
    assert errors[0][0] == "Dossier des assets invalide"
    assert called == []


def _blocking_report(*issues):
    return bs.ValidationReport(list(issues))


def _blocking_issue(code="DUPLICATE_OUTPUT_TARGET", entity="book", message="collision"):
    return bs.ValidationIssue("blocking", code, entity, "id", "slug", message)


def test_gui_message_collision_livres_explicite_simple():
    books = pd.DataFrame([
        {
            "titre_norm": "Premier ouvrage",
            "id13": "9782877750001",
            "_source_slug_raw": "meme-slug",
            "_source_slug": "meme-slug",
            "slug": "meme-slug",
        },
        {
            "titre_norm": "Second ouvrage",
            "id13": "9791024000008",
            "_source_slug_raw": "meme-slug",
            "_source_slug": "meme-slug",
            "slug": "meme-slug",
        },
    ], index=[197, 457])
    title, message, log_message = format_blocking_validation_message(
        _blocking_report(_blocking_issue()), books
    )
    assert title == "URLs de livres en conflit"
    assert "livres/meme-slug.html" in message
    assert "ligne 199" in message
    assert "ligne 459" in message
    assert "Premier ouvrage" in message
    assert "Second ouvrage" in message
    assert "ISBN 9782877750001" in message
    assert "ISBN 9791024000008" in message
    assert "colonne \u00ab slug \u00bb" in message
    assert "modifiez" in message
    assert "n'est pas" not in message
    assert "livres/meme-slug.html" in log_message
    assert "g\u00e9n\u00e9ration est arr\u00eat\u00e9e" in message
    assert "valeur diff\u00e9rente" in message
    assert "Collisions d'URL de livres d\u00e9tect\u00e9es" in log_message


def test_gui_message_collision_livres_slugs_calcules():
    books = pd.DataFrame([
        {"titre_norm": "Premier", "id13": "", "_source_slug_raw": "", "_source_slug": "", "slug": "meme-slug"},
        {"titre_norm": "Second", "id13": "", "_source_slug_raw": "", "_source_slug": "", "slug": "meme-slug"},
    ], index=[0, 1])
    _, message, _ = format_blocking_validation_message(_blocking_report(_blocking_issue()), books)
    assert "slug Excel : (vide \u2014 URL calcul\u00e9e automatiquement)" in message


def test_gui_message_collision_livres_sources_brutes_depuis_load_books(tmp_path):
    wb = _workbook(
        tmp_path / "raw-slugs.xlsx",
        book_rows=[
            {"slug": "Mon Livre", "id13": "9782877750001", "titre_norm": "Premier"},
            {"slug": "mon-livre", "id13": "9782877750002", "titre_norm": "Second"},
        ],
    )
    books = _loaded_books(wb)
    assert books["_source_slug_raw"].tolist() == ["Mon Livre", "mon-livre"]
    assert books["_source_slug"].tolist() == ["mon-livre", "mon-livre"]
    assert books["slug"].tolist() == ["mon-livre", "mon-livre"]

    collisions = find_book_url_collisions(books)
    assert len(collisions) == 1
    _, message, _ = format_blocking_validation_message(_blocking_report(_blocking_issue()), books)
    assert message.count("livres/mon-livre.html") == 1
    assert "slug Excel : Mon Livre" in message
    assert "slug Excel : mon-livre" in message
    assert "colonne \u00ab slug \u00bb" in message


def test_gui_message_collision_livres_fallback_ancien_dataframe_sans_slug_raw():
    books = pd.DataFrame([
        {"titre_norm": "Premier", "id13": "", "_source_slug": "Mon Livre", "slug": "mon-livre"},
        {"titre_norm": "Second", "id13": "", "_source_slug": "mon-livre", "slug": "mon-livre"},
    ])
    _, message, _ = format_blocking_validation_message(_blocking_report(_blocking_issue()), books)
    assert "slug Excel : Mon Livre" in message
    assert "slug Excel : mon-livre" in message


def test_gui_message_collision_livres_trois_ouvrages():
    books = pd.DataFrame([
        {"titre_norm": "Premier", "id13": "", "_source_slug_raw": "meme-slug", "slug": "meme-slug"},
        {"titre_norm": "Deuxieme", "id13": "", "_source_slug_raw": "meme-slug", "slug": "meme-slug"},
        {"titre_norm": "Troisieme", "id13": "", "_source_slug_raw": "meme-slug", "slug": "meme-slug"},
    ])
    _, message, _ = format_blocking_validation_message(_blocking_report(_blocking_issue()), books)
    assert "Deux ouvrages actifs" not in message
    assert "Plusieurs ouvrages actifs produisent la m\u00eame URL" in message


def test_gui_message_collision_livres_limite_dialogue_et_detail_journal():
    rows = []
    for i in range(6):
        rows.extend([
            {"titre_norm": f"Premier {i}", "id13": "", "_source_slug": f"slug-{i}", "slug": f"slug-{i}"},
            {"titre_norm": f"Second {i}", "id13": "", "_source_slug": f"slug-{i}", "slug": f"slug-{i}"},
        ])
    books = pd.DataFrame(rows)
    title, message, log_message = format_blocking_validation_message(
        _blocking_report(_blocking_issue()), books
    )
    assert title == "URLs de livres en conflit"
    assert message.count("livres/slug-") == 5
    assert "Plusieurs ouvrages actifs produisent des URL identiques" in message
    assert "... et 1 autre(s) URL en conflit." in message
    assert log_message.count("livres/slug-") == 6


def test_gui_message_blocage_generique_sans_collision_livre():
    report = _blocking_report(
        bs.ValidationIssue("blocking", "OUTPUT_PATH_NOT_DIRECTORY", "site", "dist", "out_dir", "Sortie invalide.")
    )
    title, message, _ = format_blocking_validation_message(report, pd.DataFrame())
    assert title == "Blocage de validation"
    assert "OUTPUT_PATH_NOT_DIRECTORY" in message
    assert "Sortie invalide." in message


def test_gui_message_collision_livres_et_autre_blocage():
    books = pd.DataFrame([
        {"titre_norm": "Premier", "id13": "", "_source_slug": "meme-slug", "slug": "meme-slug"},
        {"titre_norm": "Second", "id13": "", "_source_slug": "meme-slug", "slug": "meme-slug"},
    ])
    report = _blocking_report(
        _blocking_issue(),
        bs.ValidationIssue("blocking", "OUTPUT_PATH_NOT_DIRECTORY", "site", "dist", "out_dir", "Sortie invalide."),
    )
    title, message, _ = format_blocking_validation_message(report, books)
    assert title == "URLs de livres en conflit"
    assert "livres/meme-slug.html" in message
    assert "Autres probl\u00e8mes bloquants" in message
    assert "OUTPUT_PATH_NOT_DIRECTORY" in message


def _workbook(path: Path, title: str = "Un livre", id13: str = "9782877750001",
              ftp: bool = False, book_rows: list[dict] | None = None,
              pages_rows: list[dict] | None = None) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFIG"
    ws.append(["key", "value"])
    for key, value in [
        ("site_title", "Site test"),
        ("books_sheet", "CATALOGUE"),
        ("pages_sheet", "PAGES"),
        ("collections_sheet", "COLLECTIONS"),
        ("revues_sheet", "REVUES"),
        ("contacts_sheet", "CONTACTS"),
    ]:
        ws.append([key, value])
    if ftp:
        for key, value in [
            ("ftp_host", "ftp.example.test"),
            ("ftp_user", "user"),
            ("ftp_password", "secret"),
            ("ftp_remote_dir", "/www"),
        ]:
            ws.append([key, value])

    books = wb.create_sheet("CATALOGUE")
    books.append([
        "id13", "slug", "titre_norm", "sous_titre_norm", "credit_ligne",
        "collection", "collection_id", "date_parution_norm", "format_site",
        "price", "availability", "cover_file", "Description courte",
        "Description longue", "order_url", "openedition_url", "active_site",
    ])
    if book_rows is None:
        book_rows = [{"id13": id13, "slug": "un-livre", "titre_norm": title}]
    for row in book_rows:
        books.append([
            row.get("id13", id13),
            row.get("slug", "un-livre"),
            row.get("titre_norm", title),
            row.get("sous_titre_norm", ""),
            row.get("credit_ligne", ""),
            row.get("collection", "Essais"),
            row.get("collection_id", "col-essais"),
            row.get("date_parution_norm", "2026-01-01"),
            row.get("format_site", "Broche"),
            row.get("price", "12"),
            row.get("availability", "Disponible"),
            row.get("cover_file", ""),
            row.get("Description courte", "Resume"),
            row.get("Description longue", ""),
            row.get("order_url", ""),
            row.get("openedition_url", ""),
            row.get("active_site", 1),
        ])

    pages = wb.create_sheet("PAGES")
    pages.append(["slug", "title", "content_md", "is_published"])
    if pages_rows is None:
        pages_rows = [{"slug": "presentation", "title": "Presentation", "content_md": "Texte", "is_published": 1}]
    for row in pages_rows:
        pages.append([
            row.get("slug", ""),
            row.get("title", ""),
            row.get("content_md", ""),
            row.get("is_published", 1),
        ])

    collections = wb.create_sheet("COLLECTIONS")
    collections.append(["collection_id", "name", "slug", "is_active"])
    collections.append(["col-essais", "Essais", "essais", 1])

    revues = wb.create_sheet("REVUES")
    revues.append(["revue_id", "title", "slug", "is_active"])

    contacts = wb.create_sheet("CONTACTS")
    contacts.append(["label", "name", "role", "email", "phone", "address", "order", "is_active"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path
