# Tests de actualites_store.py (logique Excel de l'éditeur d'actualités)
# Lancement : .venv\Scripts\python.exe -m pytest tests/ -v
#
# Tous les tests travaillent sur des classeurs temporaires générés ;
# le classeur réel n'est jamais touché.

import sys
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

import openpyxl
import pytest

from actualites_store import (
    ActualitesStore, Actu, ActuError, WorkbookLockedError,
    parse_date_fr, format_date_fr, normalize_id13,
)

FTP_SECRET = "MotDePasseTresSecret!42"


@pytest.fixture
def workbook(tmp_path) -> Path:
    """Classeur minimal : ACTUS + Master_Site + CONFIG (sensible) + feuille à formule."""
    path = tmp_path / "classeur.xlsx"
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "ACTUS"
    ws.append(["Titre", "Texte", "Date", "Image", "is_active", "id13", "Lien"])
    ws.append(["Première actu", "Texte avec <i>italique</i>.", datetime(2026, 6, 30),
               "img1.jpg", 1, 9791024017730, "https://example.org/"])
    ws.append(["Actu inactive", "Ancienne annonce.", datetime(2025, 1, 15),
               None, 0, None, None])

    ms = wb.create_sheet("Master_Site")
    ms.append(["id13", "titre_norm", "slug", "cover_file"])
    ms.append([9791024017730, "Wagner vu de Rouen", "wagner-vu-de-rouen-9791024017730",
               "9791024017730.jpg"])
    ms.append([9791024019826, "Penser l'éducation, n° 58", None, "9791024019826.jpg"])

    cfg = wb.create_sheet("CONFIG")
    cfg.append(["key", "value"])
    cfg.append(["ftp_password", FTP_SECRET])
    cfg.append(["site_title", "PURH"])

    f = wb.create_sheet("FORMULES")
    f["A1"] = "=SUM(1,2)"
    f["B1"] = 123

    wb.save(path)
    return path


def choose(answer: str):
    """Callback de conflit préprogrammé."""
    calls = []

    def cb(name: str) -> str:
        calls.append(name)
        return answer

    cb.calls = calls
    return cb


# ---------------------------------------------------------------------------
# Lecture
# ---------------------------------------------------------------------------

def test_lecture_actus_existantes(workbook):
    store = ActualitesStore(workbook)
    actus = store.list_actus()
    assert len(actus) == 2
    a = actus[0]
    assert a.row == 2
    assert a.title == "Première actu"
    assert "<i>italique</i>" in a.text  # le HTML léger n'est pas réécrit
    assert a.date == date(2026, 6, 30)
    assert a.is_active is True
    assert a.id13 == "9791024017730"
    assert a.link == "https://example.org/"
    assert actus[1].is_active is False


def test_feuille_actus_absente(tmp_path):
    p = tmp_path / "vide.xlsx"
    wb = openpyxl.Workbook()
    wb.save(p)
    with pytest.raises(ActuError, match="ACTUS"):
        ActualitesStore(p)


# ---------------------------------------------------------------------------
# Dates
# ---------------------------------------------------------------------------

def test_parse_date_fr_valide():
    assert parse_date_fr("31/12/2026") == date(2026, 12, 31)
    assert parse_date_fr("2026-12-31") == date(2026, 12, 31)
    assert parse_date_fr("") is None


def test_parse_date_fr_invalide():
    with pytest.raises(ValueError, match="JJ/MM/AAAA"):
        parse_date_fr("31/13/2026")
    with pytest.raises(ValueError):
        parse_date_fr("demain")


def test_date_ecrite_comme_vraie_date_avec_format_francais(workbook):
    store = ActualitesStore(workbook)
    a = Actu(title="Datée", text="t", date=parse_date_fr("24/10/2025"))
    row = store.save_actu(a)
    store.save_workbook()

    wb = openpyxl.load_workbook(workbook)
    cell = wb["ACTUS"].cell(row=row, column=3)
    assert isinstance(cell.value, (date, datetime))
    assert (cell.value.date() if isinstance(cell.value, datetime) else cell.value) == date(2025, 10, 24)
    assert cell.number_format == "DD/MM/YYYY"
    assert format_date_fr(date(2025, 10, 24)) == "24/10/2025"


# ---------------------------------------------------------------------------
# Ajout / modification / désactivation
# ---------------------------------------------------------------------------

def test_ajout_nouvelle_ligne(workbook):
    store = ActualitesStore(workbook)
    a = Actu(title="Nouvelle", text="Contenu", is_active=True)
    row = store.save_actu(a)
    assert row == 4  # première ligne vide après les données
    store.save_workbook()

    store2 = ActualitesStore(workbook)
    actus = store2.list_actus()
    assert len(actus) == 3
    assert actus[2].title == "Nouvelle"
    # en-têtes intacts
    wb = openpyxl.load_workbook(workbook)
    assert [c.value for c in wb["ACTUS"][1]][:7] == \
        ["Titre", "Texte", "Date", "Image", "is_active", "id13", "Lien"]


def test_modification_conserve_la_ligne(workbook):
    store = ActualitesStore(workbook)
    a = store.list_actus()[0]
    a.title = "Titre corrigé"
    row = store.save_actu(a)
    assert row == 2
    store.save_workbook()

    actus = ActualitesStore(workbook).list_actus()
    assert actus[0].row == 2
    assert actus[0].title == "Titre corrigé"
    assert actus[1].title == "Actu inactive"  # les autres lignes ne bougent pas


def test_desactivation_reactivation(workbook):
    store = ActualitesStore(workbook)
    store.set_active(2, False)
    store.save_workbook()
    assert ActualitesStore(workbook).list_actus()[0].is_active is False

    store2 = ActualitesStore(workbook)
    store2.set_active(2, True)
    store2.save_workbook()
    assert ActualitesStore(workbook).list_actus()[0].is_active is True


# ---------------------------------------------------------------------------
# ISBN / catalogue
# ---------------------------------------------------------------------------

def test_isbn_valide_trouve(workbook):
    store = ActualitesStore(workbook)
    i13, book = store.lookup_isbn("979-10-240-1773-0")
    assert i13 == "9791024017730"
    assert book is not None
    assert book.title == "Wagner vu de Rouen"
    assert book.cover_file == "9791024017730.jpg"
    assert book.slug == "wagner-vu-de-rouen-9791024017730"


def test_isbn_invalide(workbook):
    store = ActualitesStore(workbook)
    assert store.lookup_isbn("1234") == ("", None)
    assert normalize_id13("abc") == ""


def test_isbn_absent_du_catalogue(workbook):
    store = ActualitesStore(workbook)
    i13, book = store.lookup_isbn("9999999999999")
    assert i13 == "9999999999999"
    assert book is None
    issues = store.validate(Actu(title="t", text="t", id13="9999999999999"))
    assert any(i.confirmable and "catalogue" in i.message for i in issues)


def test_validation_messages(workbook):
    store = ActualitesStore(workbook)
    # ni titre ni texte
    assert any("titre ou un texte" in i.message for i in store.validate(Actu()))
    # lien invalide
    issues = store.validate(Actu(title="t", link="www.example.org"))
    assert any("http://" in i.message for i in issues)
    # image introuvable, message pour éditrice
    issues = store.validate(Actu(title="t", image="catalogue2026.jpg"))
    msg = [i.message for i in issues if "catalogue2026.jpg" in i.message]
    assert msg and "n'a pas été trouvée" in msg[0]
    assert "Traceback" not in msg[0]


# ---------------------------------------------------------------------------
# Images
# ---------------------------------------------------------------------------

def test_import_image_copie_dans_actu(workbook, tmp_path):
    store = ActualitesStore(workbook)
    src = tmp_path / "photo.jpg"
    src.write_bytes(b"JPEGDATA")
    name = store.import_image(src, choose("replace"))
    assert name == "photo.jpg"
    assert (workbook.parent / "actu" / "photo.jpg").read_bytes() == b"JPEGDATA"
    # la valeur stockée est résoluble par la logique de Cidre
    assert store.resolve_image("photo.jpg") is not None


def test_import_image_collision(workbook, tmp_path):
    store = ActualitesStore(workbook)
    (workbook.parent / "actu").mkdir()
    (workbook.parent / "actu" / "photo.jpg").write_bytes(b"ANCIEN")
    src = tmp_path / "photo.jpg"
    src.write_bytes(b"NOUVEAU")

    # conserver l'existant
    cb = choose("keep")
    assert store.import_image(src, cb) == "photo.jpg"
    assert cb.calls == ["photo.jpg"]
    assert (workbook.parent / "actu" / "photo.jpg").read_bytes() == b"ANCIEN"

    # annuler
    assert store.import_image(src, choose("cancel")) is None
    assert (workbook.parent / "actu" / "photo.jpg").read_bytes() == b"ANCIEN"

    # remplacer
    assert store.import_image(src, choose("replace")) == "photo.jpg"
    assert (workbook.parent / "actu" / "photo.jpg").read_bytes() == b"NOUVEAU"


def test_import_image_manquante_ou_mauvais_format(workbook, tmp_path):
    store = ActualitesStore(workbook)
    with pytest.raises(ActuError, match="introuvable"):
        store.import_image(tmp_path / "absente.jpg", choose("replace"))
    bad = tmp_path / "doc.pdf"
    bad.write_bytes(b"%PDF")
    with pytest.raises(ActuError, match="format"):
        store.import_image(bad, choose("replace"))


def test_use_cover_sans_copie_si_deja_resolue(workbook):
    store = ActualitesStore(workbook)
    actu_dir = workbook.parent / "actu"
    actu_dir.mkdir()
    (actu_dir / "9791024017730.jpg").write_bytes(b"COVER")
    name = store.use_cover("9791024017730.jpg", choose("replace"))
    assert name == "9791024017730.jpg"  # aucun fichier recopié : déjà trouvable


def test_use_cover_copie_depuis_covers(workbook):
    store = ActualitesStore(workbook)
    covers = workbook.parent / "covers"
    covers.mkdir()
    (covers / "9791024017730.jpg").write_bytes(b"COVER")
    name = store.use_cover("9791024017730.jpg", choose("replace"))
    assert name == "9791024017730.jpg"
    assert (workbook.parent / "actu" / "9791024017730.jpg").read_bytes() == b"COVER"


def test_use_cover_introuvable(workbook):
    store = ActualitesStore(workbook)
    with pytest.raises(ActuError, match="Choisir une image"):
        store.use_cover("inexistante.jpg", choose("replace"))


# ---------------------------------------------------------------------------
# Sécurité du classeur
# ---------------------------------------------------------------------------

def test_sauvegarde_ne_modifie_pas_les_autres_feuilles(workbook):
    store = ActualitesStore(workbook)
    store.save_actu(Actu(title="X", text="Y"))
    store.save_workbook()

    wb = openpyxl.load_workbook(workbook)
    assert wb["Master_Site"]["B2"].value == "Wagner vu de Rouen"
    assert wb["CONFIG"]["B2"].value == FTP_SECRET  # CONFIG intact
    assert wb["FORMULES"]["A1"].value == "=SUM(1,2)"  # formule préservée
    assert wb["FORMULES"]["B1"].value == 123


def test_creation_sauvegarde_horodatee(workbook):
    store = ActualitesStore(workbook)
    store.save_actu(Actu(title="X", text="Y"))
    store.save_workbook()
    backups = list(workbook.parent.glob("classeur.sauvegarde-*.xlsx"))
    assert len(backups) == 1
    # la sauvegarde est le classeur AVANT modification
    wb = openpyxl.load_workbook(backups[0])
    assert wb["ACTUS"].cell(row=4, column=1).value is None

    # une seule sauvegarde par session, même après plusieurs enregistrements
    store.save_actu(Actu(title="Z", text="Z"))
    store.save_workbook()
    assert len(list(workbook.parent.glob("classeur.sauvegarde-*.xlsx"))) == 1


def test_classeur_ouvert_dans_excel(workbook):
    store = ActualitesStore(workbook)
    lock = workbook.parent / f"~${workbook.name}"
    lock.write_bytes(b"")
    with pytest.raises(WorkbookLockedError, match="Fermez le classeur"):
        store.save_workbook()
    lock.unlink()
    store.save_workbook()  # repasse une fois le verrou levé


def test_config_jamais_exposee(workbook, capsys):
    """Aucune API publique du store ne doit laisser fuiter le contenu de CONFIG."""
    store = ActualitesStore(workbook)
    outputs = []
    outputs.append(repr(store.list_actus()))
    outputs.append(repr(store.books()))
    outputs.append(repr(store.lookup_isbn("9791024017730")))
    outputs.append(repr(store.validate(Actu(title="t", id13="9999999999999"))))
    outputs.append(store.sheet_name)
    captured = capsys.readouterr()
    blob = "\n".join(outputs) + captured.out + captured.err
    assert FTP_SECRET not in blob
    assert "ftp_password" not in blob
    # la détection du catalogue ignore explicitement CONFIG
    assert store._detect_books_sheet() == "Master_Site"


# ---------------------------------------------------------------------------
# Indépendance des champs id13 / Image / Lien (modèle fonctionnel)
# ---------------------------------------------------------------------------
# Une actualité n'est pas nécessairement la promotion d'un ouvrage :
# chaque combinaison doit s'enregistrer sans erreur ni avertissement indu.

ISBN_CATALOGUE = "9791024017730"  # présent dans la fixture Master_Site


def _roundtrip(workbook, actu: Actu) -> Actu:
    """Valide (aucun problème attendu), enregistre, relit depuis le fichier."""
    store = ActualitesStore(workbook)
    issues = store.validate(actu)
    assert issues == [], [i.message for i in issues]
    row = store.save_actu(actu)
    store.save_workbook()
    rel = [a for a in ActualitesStore(workbook).list_actus() if a.row == row]
    assert len(rel) == 1
    return rel[0]


def _with_cover(workbook) -> str:
    """Rend la couverture du livre catalogue résoluble par Cidre."""
    d = workbook.parent / "actu"
    d.mkdir(exist_ok=True)
    (d / "9791024017730.jpg").write_bytes(b"COVER")
    return "9791024017730.jpg"


def test_cas1_isbn_avec_couverture(workbook):
    cover = _with_cover(workbook)
    a = _roundtrip(workbook, Actu(title="Prix littéraire", text="t",
                                  id13=ISBN_CATALOGUE, image=cover))
    assert (a.id13, a.image) == (ISBN_CATALOGUE, cover)


def test_cas2_isbn_avec_visuel_specifique(workbook):
    d = workbook.parent / "actu"
    d.mkdir(exist_ok=True)
    (d / "remise-du-prix.jpg").write_bytes(b"PHOTO")
    a = _roundtrip(workbook, Actu(title="Remise du prix", text="t",
                                  id13=ISBN_CATALOGUE, image="remise-du-prix.jpg"))
    assert a.id13 == ISBN_CATALOGUE
    assert a.image == "remise-du-prix.jpg"  # le visuel n'est pas la couverture


def test_cas3_isbn_sans_image(workbook):
    a = _roundtrip(workbook, Actu(title="Réimpression annoncée", text="t",
                                  id13=ISBN_CATALOGUE))
    assert a.id13 == ISBN_CATALOGUE
    assert a.image == ""


def test_cas4_sans_isbn_avec_visuel(workbook):
    d = workbook.parent / "actu"
    d.mkdir(exist_ok=True)
    (d / "salon-du-livre.jpg").write_bytes(b"AFFICHE")
    a = _roundtrip(workbook, Actu(title="Salon du livre de Rouen",
                                  text="Retrouvez-nous stand 12.",
                                  image="salon-du-livre.jpg"))
    assert a.id13 == ""
    assert a.image == "salon-du-livre.jpg"


def test_cas5_sans_isbn_sans_image(workbook):
    a = _roundtrip(workbook, Actu(title="Fermeture estivale",
                                  text="Les PURH seront fermées en août."))
    assert (a.id13, a.image, a.link) == ("", "", "")


def test_cas6_isbn_et_lien_externe(workbook):
    a = _roundtrip(workbook, Actu(title="Recension dans la presse", text="t",
                                  id13=ISBN_CATALOGUE,
                                  link="https://journal.example.org/article"))
    # les deux champs coexistent : l'ISBN pilote le lien de la fiche,
    # le lien externe est affiché en plus par Cidre
    assert a.id13 == ISBN_CATALOGUE
    assert a.link == "https://journal.example.org/article"


def test_cas7_sans_isbn_avec_lien_externe(workbook):
    a = _roundtrip(workbook, Actu(title="Colloque : inscriptions ouvertes", text="t",
                                  link="https://colloque.example.org/inscription"))
    assert a.id13 == ""
    assert a.link == "https://colloque.example.org/inscription"


def test_absence_isbn_ne_genere_aucun_avertissement(workbook):
    store = ActualitesStore(workbook)
    issues = store.validate(Actu(title="Sans livre", text="t"))
    assert issues == []
    # et aucune validation "confirmable" non plus
    assert not any(i.confirmable for i in issues)


def test_seule_une_image_renseignee_mais_introuvable_bloque(workbook):
    store = ActualitesStore(workbook)
    # image absente : aucun problème
    assert store.validate(Actu(title="t", text="t")) == []
    # image renseignée mais introuvable : blocage explicite
    issues = store.validate(Actu(title="t", text="t", image="fantome.jpg"))
    assert len(issues) == 1 and not issues[0].confirmable
    assert "fantome.jpg" in issues[0].message
