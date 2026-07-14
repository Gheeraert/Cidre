import json
import re
import sys
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl
import pandas as pd

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import build_site as bs
from cidre.validation import validate_site_data
from cidre.seo import (
    PageSeo,
    absolute_public_url,
    clean_description,
    json_ld_script,
    normalize_site_url,
    robots_txt,
    seo_head_html,
)


def _workbook(path: Path, *, site_url: str = "https://example.org/purh") -> Path:
    workbook = openpyxl.Workbook()
    config = workbook.active
    config.title = "CONFIG"
    config.append(["key", "value"])
    for key, value in [
        ("site_title", "Presses test"),
        ("site_subtitle", "Catalogue de recherche"),
        ("site_url", site_url),
        ("site_description", "Une description éditoriale du site test."),
        ("social_image", "social-share.jpg"),
        ("books_sheet", "CATALOGUE"),
        ("pages_sheet", "PAGES"),
        ("collections_sheet", "COLLECTIONS"),
        ("revues_sheet", "REVUES"),
        ("contacts_sheet", "CONTACTS"),
        ("social_1_name", "Mastodon"),
        ("social_1_url", "https://social.example/@presses"),
    ]:
        config.append([key, value])

    books = workbook.create_sheet("CATALOGUE")
    books.append([
        "id13", "slug", "titre_norm", "sous_titre_norm", "credit_ligne",
        "collection", "collection_id", "date_parution_norm", "format_site",
        "price", "availability", "cover_file", "Description courte",
        "Description longue", "order_url", "openedition_url", "active_site",
    ])
    books.append([
        "9782877750001", "livre-avec-couverture", "Livre avec couverture", "Sous-titre",
        "Sous la direction de Personne", "Essais", "essais", "2026-01-02", "Broché",
        "18", "Disponible", "presente.jpg", "Description **courte** <em>utile</em>",
        "", "", "https://books.openedition.org/exemple", 1,
    ])
    books.append([
        "9782877750002", "livre-sans-couverture", "Livre sans couverture", "", "",
        "Essais", "essais", "2026", "Numérique", "12", "Disponible", "absente.jpg",
        "</script><script>alert(1)</script> Description longue", "", "", "", 1,
    ])

    pages = workbook.create_sheet("PAGES")
    pages.append(["slug", "title", "content_md", "is_published"])
    pages.append(["presentation", "Présentation", "Texte **éditorial** de la maison.", 1])

    collections = workbook.create_sheet("COLLECTIONS")
    collections.append(["collection_id", "name", "slug", "description_md", "is_active"])
    collections.append(["essais", "Essais", "essais", "Collection *Essais*.", 1])

    revues = workbook.create_sheet("REVUES")
    revues.append(["journal_id", "title", "slug", "description_md", "is_active"])
    revues.append(["revue-test", "Revue test", "revue-test", "Une revue de test.", 1])

    contacts = workbook.create_sheet("CONTACTS")
    contacts.append(["label", "name", "role", "email", "phone", "address", "order", "is_active"])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _json_ld(html: str) -> list[dict]:
    payloads = re.findall(r'<script type="application/ld\+json">(.*?)</script>', html, flags=re.S)
    return [json.loads(payload) for payload in payloads]


def test_url_publique_normalisee_et_sous_chemin():
    assert normalize_site_url(" https://example.org/purh/ ") == "https://example.org/purh"
    assert absolute_public_url("https://example.org/purh/", "livres/test.html") == "https://example.org/purh/livres/test.html"
    assert absolute_public_url("https://example.org/", "index.html") == "https://example.org"
    assert absolute_public_url("https://example.org//purh//", "livres/test.html") == "https://example.org/purh/livres/test.html"
    assert normalize_site_url("https://example.org/purh?x=1") == ""
    assert normalize_site_url("https://example.org/purh#ancre") == ""
    assert normalize_site_url("/purh") == ""


def test_description_nettoyee_tronquee_et_head_sans_url_incorrecte():
    description = clean_description("<p>Texte **avec**   espaces et <em>balises</em>.</p>")
    assert description == "Texte avec espaces et balises."
    assert clean_description("mot " * 80).endswith("…")

    cfg = bs.SiteConfig(site_title='Site "test"', site_url="")
    head = seo_head_html(cfg, 'Titre "test"', PageSeo(description='Description "utile"'))
    assert 'name="description" content="Description &quot;utile&quot;"' in head
    assert "canonical" not in head
    assert 'property="og:url"' not in head
    assert 'name="twitter:card" content="summary"' in head


def test_json_ld_ne_peut_pas_fermer_son_script():
    script = json_ld_script({"@context": "https://schema.org", "description": "</script><script>alert(1)</script>"})
    assert "</script><script>" not in script
    payload = re.search(r">(.*?)</script>$", script, flags=re.S).group(1)
    assert json.loads(payload)["description"] == "</script><script>alert(1)</script>"


def test_generation_seo_complete_et_sitemap_deterministe(tmp_path):
    workbook = _workbook(tmp_path / "site.xlsx")
    covers = tmp_path / "covers"
    covers.mkdir()
    (covers / "presente.jpg").write_bytes(b"image")
    assets = tmp_path / "assets-source"
    assets.mkdir()
    (assets / "social-share.jpg").write_bytes(b"image")
    output = tmp_path / "site-sortie"

    bs.build_site(workbook, output, covers_dir=covers, assets_dir=assets)

    home = (output / "index.html").read_text(encoding="utf-8")
    book = (output / "livres" / "livre-avec-couverture.html").read_text(encoding="utf-8")
    missing_cover_book = (output / "livres" / "livre-sans-couverture.html").read_text(encoding="utf-8")
    assert '<link rel="canonical" href="https://example.org/purh">' in home
    assert '<meta property="og:image" content="https://example.org/purh/assets/social-share.jpg">' in home
    assert (output / "assets" / "social-share.jpg").is_file()
    assert '<meta property="og:url" content="https://example.org/purh/livres/livre-avec-couverture.html">' in book
    assert '<meta property="og:image" content="https://example.org/purh/covers/presente.jpg">' in book
    assert "alt='Couverture de Livre avec couverture'" in book
    assert "absente.jpg" not in missing_cover_book
    assert "<h1>Livre avec couverture</h1>" in book
    assert "<h2 class='section-heading'>Présentation</h2>" in book
    assert "<h2>Table des matières</h2>" not in book
    assert "Presses test — Livre avec couverture" not in book
    assert "Livre avec couverture — Presses test" in book

    home_ld = _json_ld(home)
    assert {item["@type"] for item in home_ld} == {"Organization", "WebSite"}
    book_ld = _json_ld(book)[0]
    assert book_ld["@type"] == "Book"
    assert book_ld["isbn"] == "9782877750001"
    assert book_ld["datePublished"] == "2026-01-02"
    assert "author" not in book_ld
    assert book_ld["image"] == "https://example.org/purh/covers/presente.jpg"
    assert "image" not in _json_ld(missing_cover_book)[0]

    sitemap = output / "sitemap.xml"
    root = ET.parse(sitemap).getroot()
    namespace = "{http://www.sitemaps.org/schemas/sitemap/0.9}"
    urls = [loc.text for loc in root.findall(f"{namespace}url/{namespace}loc")]
    assert urls == sorted(urls)
    assert len(urls) == len(set(urls))
    assert "https://example.org/purh" in urls
    assert "https://example.org/purh/livres/livre-avec-couverture.html" in urls
    assert "https://example.org/purh/collections/essais.html" in urls
    assert "https://example.org/purh/revues/revue-test.html" in urls
    assert "https://example.org/purh/presentation.html" in urls
    assert not any(url.endswith((".json", ".csv")) or "#" in url for url in urls)
    assert "priority" not in sitemap.read_text(encoding="utf-8")
    assert "changefreq" not in sitemap.read_text(encoding="utf-8")
    assert "lastmod" not in sitemap.read_text(encoding="utf-8")
    assert (output / "robots.txt").read_text(encoding="utf-8") == (
        "User-agent: *\nAllow: /\n\nSitemap: https://example.org/purh/sitemap.xml\n"
    )


def test_regeneration_supprime_les_fichiers_seo_perimes(tmp_path):
    first = _workbook(tmp_path / "first.xlsx", site_url="https://example.org/purh")
    second = _workbook(tmp_path / "second.xlsx", site_url="")
    output = tmp_path / "site-sortie"
    bs.build_site(first, output, covers_dir=None)
    assert (output / "sitemap.xml").exists()

    bs.build_site(second, output, covers_dir=None)
    assert not (output / "sitemap.xml").exists()
    assert (output / "robots.txt").read_text(encoding="utf-8") == "User-agent: *\nAllow: /\n"


def test_robots_sans_site_url():
    assert robots_txt("") == "User-agent: *\nAllow: /\n"


def test_validation_signale_site_url_absente_ou_invalide():
    books = pd.DataFrame()
    absent = validate_site_data(books=books, cfg=bs.SiteConfig(site_url=""))
    invalid = validate_site_data(books=books, cfg=bs.SiteConfig(site_url="https://example.org/?q=1"))
    assert [(issue.code, issue.severity) for issue in absent.issues] == [("SITE_URL_MISSING", "warning")]
    assert [(issue.code, issue.severity) for issue in invalid.issues] == [("SITE_URL_INVALID", "warning")]


def test_gabarit_officiel_reconnait_les_cles_seo():
    template = ROOT / "gabarit" / "purh_site_excel_gabarit.xlsx"
    with pd.ExcelFile(template) as workbook:
        cfg = bs.load_config(workbook, "CONFIG")
    assert cfg.site_url == ""
    assert cfg.site_description == ""
    assert cfg.social_image == ""
