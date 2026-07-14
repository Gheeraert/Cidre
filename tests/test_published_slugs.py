import json
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

import build_site as bs
import gui_tk
from gui_tk import App, format_slug_change_message, slug_correction_text
from cidre.published_slugs import (
    compare_published_book_slugs,
    published_slug_issues,
)


class _FakeVar:
    def __init__(self, value):
        self.value = value

    def get(self):
        return self.value

    def set(self, value):
        self.value = value


class _FakeButton:
    def config(self, **kwargs):
        pass


class _ImmediateThread:
    def __init__(self, target, daemon=None):
        self.target = target

    def start(self):
        self.target()


def _catalogue(path: Path, rows) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(rows, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def _books(rows) -> pd.DataFrame:
    return pd.DataFrame(rows)


def _workbook(path: Path, rows) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONFIG"
    ws.append(["key", "value"])
    for key, value in [
        ("site_title", "Site test"),
        ("books_sheet", "CATALOGUE"),
        ("pages_sheet", "PAGES"),
        ("collections_sheet", "COLLECTIONS"),
        ("revues_sheet", "REVUES"),
        ("contacts_sheet", "CONTACTS"),
    ]:
        ws.append([key, value])
    books = wb.create_sheet("CATALOGUE")
    books.append([
        "id13", "slug", "titre_norm", "sous_titre_norm", "credit_ligne",
        "collection", "collection_id", "date_parution_norm", "format_site",
        "price", "availability", "cover_file", "Description courte",
        "Description longue", "order_url", "openedition_url", "active_site",
    ])
    for row in rows:
        books.append([
            row.get("id13", "9782877750001"),
            row.get("slug", "un-livre"),
            row.get("titre_norm", "Un livre"),
            row.get("collection", "Essais"),
            row.get("collection_id", "essais"),
            "",
            "",
            "2026-01-01",
            "Broche",
            "12",
            "Disponible",
            "",
            "Résumé",
            "",
            "",
            "",
            row.get("active_site", 1),
        ])
    collections = wb.create_sheet("COLLECTIONS")
    collections.append(["collection_id", "name", "slug", "description_md", "is_active"])
    collections.append(["essais", "Essais", "essais", "", 1])
    wb.save(path)
    return path


def _snapshot(root: Path) -> dict[str, bytes]:
    return {
        str(path.relative_to(root)).replace("\\", "/"): path.read_bytes()
        for path in root.rglob("*")
        if path.is_file()
    }


def test_compare_absent_empty_and_same_slugs(tmp_path):
    books = _books([{"id13": "978-2-87775-000-1", "slug": "slug-actuel", "titre_norm": "Livre"}])
    absent = compare_published_book_slugs(tmp_path / "absent.json", books)
    assert absent.changes == []
    assert absent.problems == []

    empty = compare_published_book_slugs(_catalogue(tmp_path / "empty.json", []), books)
    assert empty.changes == []
    assert empty.problems == []

    same = compare_published_book_slugs(
        _catalogue(tmp_path / "catalogue.json", [{"id13": "9782877750001", "slug": "slug-actuel"}]),
        books,
    )
    assert same.changes == []
    assert same.problems == []


def test_compare_slug_modifie_et_proposition_ancien_slug(tmp_path):
    books = _books([{"id13": "9782877750001", "slug": "nouveau-slug", "titre_norm": "Livre"}])
    comparison = compare_published_book_slugs(
        _catalogue(tmp_path / "catalogue.json", [{"id13": "978-2-87775-000-1", "slug": "ancien-slug"}]),
        books,
    )

    assert len(comparison.changes) == 1
    change = comparison.changes[0]
    assert change.id13 == "9782877750001"
    assert change.published_slug == "ancien-slug"
    assert change.current_slug == "nouveau-slug"
    assert change.recommended_slug == "ancien-slug"
    issue = published_slug_issues(comparison)[0]
    assert issue.code == "BOOK_SLUG_CHANGED"
    assert issue.severity == "alert"
    assert "Slug recommandé : ancien-slug" in issue.message


def test_compare_plusieurs_modifies_et_isbn_ignores(tmp_path):
    books = _books([
        {"id13": "9782877750001", "slug": "new-1", "titre_norm": "Un"},
        {"id13": "9782877750002", "slug": "new-2", "titre_norm": "Deux"},
        {"id13": "9782877750003", "slug": "new-3", "titre_norm": "Nouveau"},
        {"id13": "", "slug": "sans-isbn", "titre_norm": "Sans ISBN"},
    ])
    comparison = compare_published_book_slugs(
        _catalogue(tmp_path / "catalogue.json", [
            {"id13": "9782877750001", "slug": "old-1"},
            {"id13": "9782877750002", "slug": "old-2"},
            {"id13": "9782877759999", "slug": "supprime"},
            {"id13": "", "slug": "ignore"},
        ]),
        books,
    )

    assert [c.id13 for c in comparison.changes] == ["9782877750001", "9782877750002"]
    assert [c.recommended_slug for c in comparison.changes] == ["old-1", "old-2"]


@pytest.mark.parametrize("content", ["{", "{}", "[1]"])
def test_compare_catalogue_invalide_ou_structure_invalide(tmp_path, content):
    path = tmp_path / "catalogue.json"
    path.write_text(content, encoding="utf-8")
    comparison = compare_published_book_slugs(path, _books([{"id13": "9782877750001", "slug": "x"}]))

    assert comparison.changes == []
    assert comparison.problems
    assert comparison.problems[0].code == "PUBLISHED_CATALOGUE_UNREADABLE"
    assert published_slug_issues(comparison)[0].severity == "alert"


def test_compare_catalogue_non_utf8_devient_alerte(tmp_path):
    path = tmp_path / "catalogue.json"
    path.write_bytes(b'[{"id13": "9782877750001", "slug": "\xff"}]')

    comparison = compare_published_book_slugs(path, _books([{"id13": "9782877750001", "slug": "x"}]))

    assert comparison.changes == []
    assert comparison.problems[0].code == "PUBLISHED_CATALOGUE_UNREADABLE"
    assert published_slug_issues(comparison)[0].code == "PUBLISHED_CATALOGUE_UNREADABLE"


def test_compare_ancien_catalogue_isbn_ambigu_et_slug_vide(tmp_path):
    comparison = compare_published_book_slugs(
        _catalogue(tmp_path / "catalogue.json", [
            {"id13": "9782877750001", "slug": "a"},
            {"id13": "9782877750001", "slug": "b"},
            {"id13": "9782877750002", "slug": ""},
        ]),
        _books([
            {"id13": "9782877750001", "slug": "a"},
            {"id13": "9782877750002", "slug": "x"},
        ]),
    )

    assert {p.code for p in comparison.problems} == {
        "PUBLISHED_CATALOGUE_AMBIGUOUS_ID13",
        "PUBLISHED_BOOK_SLUG_MISSING",
    }


def test_correction_copie_slug_selectionne_et_toutes_corrections(tmp_path):
    comparison = compare_published_book_slugs(
        _catalogue(tmp_path / "catalogue.json", [
            {"id13": "9782877750001", "slug": "ancien-1"},
            {"id13": "9782877750002", "slug": "ancien-2"},
        ]),
        _books([
            {"id13": "9782877750001", "slug": "nouveau-1", "titre_norm": "Un"},
            {"id13": "9782877750002", "slug": "nouveau-2", "titre_norm": "Deux"},
        ]),
    )

    assert comparison.changes[0].recommended_slug == "ancien-1"
    assert slug_correction_text(comparison.changes) == (
        "ISBN\tslug recommandé\n"
        "9782877750001\tancien-1\n"
        "9782877750002\tancien-2"
    )
    assert "Slug recommandé à recopier dans l'Excel :\nancien-1" in format_slug_change_message(comparison.changes[0])


def test_build_site_slug_change_sans_force_ne_modifie_pas_sortie_et_pas_ftp(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx", [{"id13": "9782877750001", "slug": "nouveau", "titre_norm": "Livre"}])
    out = tmp_path / "site-sortie"
    out.mkdir()
    _catalogue(out / "catalogue.json", [{"id13": "9782877750001", "slug": "ancien"}])
    (out / "index.html").write_text("ancien site", encoding="utf-8")
    before = _snapshot(out)
    ftp_calls = []
    monkeypatch.setattr("cidre.orchestrator.publish_ftp", lambda *a, **k: ftp_calls.append((a, k)))

    with pytest.raises(bs.ValidationAlertError) as exc:
        bs.build_site(wb, out, covers_dir=None, force_alerts=False, publish=True)

    assert any(i.code == "BOOK_SLUG_CHANGED" for i in exc.value.report.alerts)
    assert _snapshot(out) == before
    assert ftp_calls == []
    assert not list(tmp_path.glob(".site-sortie.build-*"))
    assert not list(tmp_path.glob(".site-sortie.backup-*"))


def test_build_site_slug_change_avec_force_genere(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx", [{"id13": "9782877750001", "slug": "nouveau", "titre_norm": "Livre"}])
    out = tmp_path / "site-sortie"
    out.mkdir()
    _catalogue(out / "catalogue.json", [{"id13": "9782877750001", "slug": "ancien"}])

    report = bs.build_site(wb, out, covers_dir=None, force_alerts=True)

    assert any(i.code == "BOOK_SLUG_CHANGED" for i in report.alerts)
    assert (out / "livres" / "nouveau.html").exists()


def test_cli_alerte_stabilite_ne_signale_pas_ancien_validation_csv(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx", [{"id13": "9782877750001", "slug": "nouveau", "titre_norm": "Livre"}])
    out = tmp_path / "site-sortie"
    out.mkdir()
    _catalogue(out / "catalogue.json", [{"id13": "9782877750001", "slug": "ancien"}])
    (out / "validation.csv").write_text("ancien rapport", encoding="utf-8")

    result = subprocess.run(
        [sys.executable, "build_site.py", "--excel", str(wb), "--out", str(out)],
        cwd=Path(__file__).resolve().parents[1],
        text=True,
        capture_output=True,
    )

    assert result.returncode == 4
    assert "BOOK_SLUG_CHANGED" in result.stderr
    assert "Rapport écrit" not in result.stderr
    assert (out / "validation.csv").read_text(encoding="utf-8") == "ancien rapport"


def test_validate_only_compare_avant_remplacement_catalogue(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx", [{"id13": "9782877750001", "slug": "nouveau", "titre_norm": "Livre"}])
    out = tmp_path / "site-sortie"
    out.mkdir()
    _catalogue(out / "catalogue.json", [{"id13": "9782877750001", "slug": "ancien"}])

    with pytest.raises(bs.ValidationAlertError):
        bs.build_site(wb, out, covers_dir=None, validate_only=True, force_alerts=False)
    assert json.loads((out / "catalogue.json").read_text(encoding="utf-8"))[0]["slug"] == "ancien"

    report = bs.build_site(wb, out, covers_dir=None, validate_only=True, force_alerts=True)
    assert any(i.code == "BOOK_SLUG_CHANGED" for i in report.alerts)
    assert "BOOK_SLUG_CHANGED" in (out / "validation.csv").read_text(encoding="utf-8")
    assert json.loads((out / "catalogue.json").read_text(encoding="utf-8"))[0]["slug"] == "nouveau"


def test_cli_slug_change_sans_et_avec_force(tmp_path):
    wb = _workbook(tmp_path / "site.xlsx", [{"id13": "9782877750001", "slug": "nouveau", "titre_norm": "Livre"}])
    out = tmp_path / "site-sortie"
    out.mkdir()
    _catalogue(out / "catalogue.json", [{"id13": "9782877750001", "slug": "ancien"}])
    before = _snapshot(out)

    cmd = [sys.executable, "build_site.py", "--excel", str(wb), "--out", str(out)]
    result = subprocess.run(cmd, cwd=Path(__file__).resolve().parents[1], text=True, capture_output=True)
    assert result.returncode == 4
    assert "BOOK_SLUG_CHANGED" in result.stderr
    assert "Slug publié : ancien" in result.stderr
    assert "Slug demandé : nouveau" in result.stderr
    assert "Slug recommandé : ancien" in result.stderr
    assert _snapshot(out) == before

    result_force = subprocess.run(cmd + ["--force"], cwd=Path(__file__).resolve().parents[1], text=True, capture_output=True)
    assert result_force.returncode == 0
    assert (out / "livres" / "nouveau.html").exists()


def test_gui_annulation_slug_change_ne_lance_pas_generation(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx", [{"id13": "9782877750001", "slug": "nouveau", "titre_norm": "Livre"}])
    out = tmp_path / "site-sortie"
    out.mkdir()
    _catalogue(out / "catalogue.json", [{"id13": "9782877750001", "slug": "ancien"}])
    before = _snapshot(out)
    build_calls = []
    confirm_calls = []

    app = object.__new__(App)
    app.var_excel = _FakeVar(str(wb))
    app.var_out = _FakeVar(str(out))
    app.var_covers = _FakeVar("")
    app.var_assets = _FakeVar("")
    app.var_validate_only = _FakeVar(False)
    app.var_publish_ftp = _FakeVar(False)
    app.var_export_onix = _FakeVar(True)
    app.var_start_server = _FakeVar(False)
    app.var_port = _FakeVar(8000)
    app.btn_run = _FakeButton()
    app.log = lambda msg: None
    app._read_cfg_from_excel = lambda excel: bs.load_config(pd.ExcelFile(excel), "CONFIG")

    monkeypatch.setattr(gui_tk, "confirm_slug_changes", lambda parent, changes: confirm_calls.append(changes) or False)
    monkeypatch.setattr(gui_tk, "build_site", lambda **kwargs: build_calls.append(kwargs))

    app.run_build()

    assert len(confirm_calls) == 1
    assert build_calls == []
    assert _snapshot(out) == before


@pytest.mark.parametrize("catalogue_rows, expected_code", [
    (None, "PUBLISHED_CATALOGUE_UNREADABLE"),
    ([
        {"id13": "9782877750001", "slug": "a"},
        {"id13": "9782877750001", "slug": "b"},
    ], "PUBLISHED_CATALOGUE_AMBIGUOUS_ID13"),
    ([{"id13": "9782877750001", "slug": ""}], "PUBLISHED_BOOK_SLUG_MISSING"),
])
def test_gui_refus_alerte_stabilite_necrit_rien(tmp_path, monkeypatch, catalogue_rows, expected_code):
    wb = _workbook(tmp_path / "site.xlsx", [{"id13": "9782877750001", "slug": "nouveau", "titre_norm": "Livre"}])
    out = tmp_path / "site-sortie"
    out.mkdir()
    if catalogue_rows is None:
        (out / "catalogue.json").write_text("{", encoding="utf-8")
    else:
        _catalogue(out / "catalogue.json", catalogue_rows)
    (out / "validation.csv").write_text("ancien rapport", encoding="utf-8")
    (out / "index.html").write_text("ancien site", encoding="utf-8")
    before = _snapshot(out)
    build_calls = []
    prompts = []

    app = object.__new__(App)
    app.var_excel = _FakeVar(str(wb))
    app.var_out = _FakeVar(str(out))
    app.var_covers = _FakeVar("")
    app.var_assets = _FakeVar("")
    app.var_validate_only = _FakeVar(False)
    app.var_publish_ftp = _FakeVar(True)
    app.var_export_onix = _FakeVar(True)
    app.var_start_server = _FakeVar(False)
    app.var_port = _FakeVar(8000)
    app.log = lambda msg: None
    app._read_cfg_from_excel = lambda excel: bs.load_config(pd.ExcelFile(excel), "CONFIG")

    def refuse(title, message):
        prompts.append((title, message))
        return False

    monkeypatch.setattr(gui_tk.messagebox, "askyesno", refuse)
    monkeypatch.setattr(gui_tk.messagebox, "showerror", lambda *args, **kwargs: None)
    monkeypatch.setattr(gui_tk, "build_site", lambda **kwargs: build_calls.append(kwargs))

    app.run_build()

    assert prompts
    assert expected_code in prompts[0][1]
    assert build_calls == []
    assert _snapshot(out) == before
    assert not list(tmp_path.glob(".site-sortie.build-*"))
    assert not list(tmp_path.glob(".site-sortie.backup-*"))


def test_gui_slug_change_force_transmet_generation(tmp_path, monkeypatch):
    wb = _workbook(tmp_path / "site.xlsx", [{"id13": "9782877750001", "slug": "nouveau", "titre_norm": "Livre"}])
    out = tmp_path / "site-sortie"
    out.mkdir()
    _catalogue(out / "catalogue.json", [{"id13": "9782877750001", "slug": "ancien"}])
    captured = []

    app = object.__new__(App)
    app.var_excel = _FakeVar(str(wb))
    app.var_out = _FakeVar(str(out))
    app.var_covers = _FakeVar("")
    app.var_assets = _FakeVar("")
    app.var_validate_only = _FakeVar(False)
    app.var_publish_ftp = _FakeVar(False)
    app.var_export_onix = _FakeVar(False)
    app.var_start_server = _FakeVar(False)
    app.var_port = _FakeVar(8000)
    app.btn_run = _FakeButton()
    app.log = lambda msg: None
    app.after = lambda delay, func=None: func() if func else None
    app.stop_server = lambda: None
    app._preview_server = None
    app._read_cfg_from_excel = lambda excel: bs.load_config(pd.ExcelFile(excel), "CONFIG")

    monkeypatch.setattr(gui_tk, "confirm_slug_changes", lambda parent, changes: True)
    monkeypatch.setattr(gui_tk.messagebox, "askyesno", lambda *args, **kwargs: True)
    monkeypatch.setattr(gui_tk.messagebox, "showerror", lambda *args, **kwargs: None)
    monkeypatch.setattr(gui_tk.threading, "Thread", _ImmediateThread)
    monkeypatch.setattr(gui_tk, "build_site", lambda **kwargs: captured.append(kwargs))

    app.run_build()

    assert captured
    assert captured[0]["force_alerts"] is True
